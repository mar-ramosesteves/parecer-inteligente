from flask import Flask, request, jsonify
from flask_cors import CORS
import os
import json
import re
import csv
import hmac
from datetime import datetime
import matplotlib.pyplot as plt
import base64
import io
import numpy as np
import requests
from openpyxl import load_workbook
from openai import OpenAI
from leadertrack_devolutivas import (
    archetype_summary,
    build_diagnostic_prompt,
    build_empty_devolutiva,
    build_history_event,
    build_integrated_plan_prompt,
    build_integrated_week1_prompt,
    build_performance_goal_suggestion,
    build_weekly_prompt,
    filter_gaps,
    low_reference_affirmations,
    microenvironment_affirmations,
    parse_json_response,
    slug,
)
from leadertrack_organizacional import (
    build_organizational_feedback_prompt,
    validate_organizational_package,
)
from leadertrack_snapshot_preview import calcular_saude_emocional_dashboard
from leadertrack_executive_snapshots import (
    SNAPSHOT_SCHEMA_VERSION,
    build_scope_snapshot,
    group_source_rows_by_company,
    snapshot_for_frontend,
    snapshot_matches_context,
)

app = Flask(__name__)
CORS(app, supports_credentials=True, resources={r"/*": {"origins": "https://gestor.thehrkey.tech"}})

ALLOWED_CORS_ORIGINS = {
    "https://gestor.thehrkey.tech",
}


@app.after_request
def add_cors_headers(response):
    origin = request.headers.get("Origin")
    if origin in ALLOWED_CORS_ORIGINS:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
        response.headers["Vary"] = "Origin"
    return response

SUPABASE_REST_URL = os.getenv("SUPABASE_REST_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
LEADERTRACK_SNAPSHOT_ADMIN_KEY = os.getenv("LEADERTRACK_SNAPSHOT_ADMIN_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
openai_client = OpenAI(api_key=OPENAI_API_KEY)


def bool_param(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in ("1", "true", "sim", "yes", "y")


def leadertrack_gap_id(gap):
    return f"{gap.get('questao')}_{slug(gap.get('dimensao'))}_{slug(gap.get('subdimensao'))}"


def selecionar_gap_leadertrack(todas_afirmacoes, dados):
    gap_enviado = dados.get("gap")
    if isinstance(gap_enviado, dict) and gap_enviado.get("questao"):
        return gap_enviado

    gap_id = str(dados.get("gapId") or dados.get("gap_id") or "").strip()
    questao = str(dados.get("questao") or "").strip().lower()
    dimensao = str(dados.get("dimensao") or "").strip().lower()
    subdimensao = str(dados.get("subdimensao") or "").strip().lower()

    for gap in todas_afirmacoes:
        if gap_id and leadertrack_gap_id(gap) == gap_id:
            return gap
        if questao and str(gap.get("questao") or "").strip().lower() == questao:
            if dimensao and str(gap.get("dimensao") or "").strip().lower() != dimensao:
                continue
            if subdimensao and str(gap.get("subdimensao") or "").strip().lower() != subdimensao:
                continue
            return gap
    return None


def intervalo_etapa_leadertrack(etapa, dados):
    etapa = str(etapa or "diagnostico").strip().lower()
    if dados.get("semanaInicio") and dados.get("semanaFim"):
        return int(dados["semanaInicio"]), int(dados["semanaFim"])
    mapa = {
        "semanas_1_4": (1, 4),
        "1_4": (1, 4),
        "semanas_5_8": (5, 8),
        "5_8": (5, 8),
        "semanas_9_12": (9, 12),
        "9_12": (9, 12),
    }
    if etapa in mapa:
        return mapa[etapa]
    return None


def carregar_prompt_leadertrack():
    """
    Carrega o prompt base do Assistente Inteligente Leadertrack.

    Este arquivo contém as regras que impedem a IA de sair do contexto
    do método Leadertrack.
    """
    try:
        with open("prompt_leadertrack_ia.txt", "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return """
ERRO: O arquivo prompt_leadertrack_ia.txt não foi encontrado.
Verifique se ele está no mesmo nível do app.py.
"""
    except Exception as e:
        return f"""
ERRO: Não foi possível carregar o prompt Leadertrack.
Detalhe técnico: {str(e)}
"""

def gerar_resposta_ia_leadertrack(
    pergunta,
    prompt_base,
    empresa,
    codrodada,
    email_lider,
    pagina_atual,
    url_atual,    
    dados_arquetipos_comparativo,
    dados_arquetipos_analitico,
    guia_arquetipos,
    dados_microambiente_analitico,
    dados_microambiente_subdimensao,
    dados_microambiente_termometro_gaps,
    dados_microambiente_waterfall_gaps,
    guia_microambiente
):
    """
    Gera uma resposta do Assistente Leadertrack com base apenas nos dados fornecidos.
    """

    contexto_leadertrack = {
        "empresa": empresa,
        "codrodada": codrodada,
        "email_lider": email_lider,
        "pagina_atual": pagina_atual,
        "url_atual": url_atual,
        "pergunta_usuario": pergunta,        
        "dados_disponiveis": {
            "arquetipos_grafico_comparativo": dados_arquetipos_comparativo,
            "arquetipos_analitico": dados_arquetipos_analitico,
            "arquetipos_parecer_ia_guia": guia_arquetipos,
            "microambiente_analitico": dados_microambiente_analitico,
            "microambiente_grafico_mediaequipe_subdimensao": dados_microambiente_subdimensao,
            "microambiente_termometro_gaps": dados_microambiente_termometro_gaps,
            "microambiente_waterfall_gaps": dados_microambiente_waterfall_gaps,
            "microambiente_parecer_ia_guia": guia_microambiente
        }
    }

    resposta = openai_client.chat.completions.create(
        model="gpt-4.1-mini",
        temperature=0.2,
        messages=[
            {
                "role": "system",
                "content": prompt_base
            },
            {
                "role": "user",
                "content": (
                    "Responda à pergunta do usuário usando exclusivamente o contexto JSON abaixo. "
                    "Não invente dados. Não use teorias externas. "
                    "Se algum dado não estiver disponível, informe a limitação.\n\n"
                    f"CONTEXTO LEADERTRACK:\n{json.dumps(contexto_leadertrack, ensure_ascii=False)}"
                )
            }
        ]
    )

    return resposta.choices[0].message.content


def gerar_resposta_ia_leadertrack_enxuta(pergunta, prompt_base, model="gpt-4.1-mini", max_tokens=3500, timeout=25, temperature=0.2):
    """
    Chamada reduzida para etapas que ja enviam o contexto estruturado no proprio prompt.
    Evita carregar graficos e relatorios inteiros quando a tela pede apenas uma semana integrada.
    """
    resposta = openai_client.chat.completions.create(
        model=model,
        temperature=temperature,
        max_tokens=max_tokens,
        response_format={"type": "json_object"},
        timeout=timeout,
        messages=[
            {
                "role": "system",
                "content": prompt_base,
            },
            {
                "role": "user",
                "content": pergunta,
            },
        ],
    )
    return resposta.choices[0].message.content


TERMOS_EXECUTIVOS_NAO_MEDIDOS = {
    "alta rotatividade": "risco organizacional nao medido nesta base",
    "retenção de talentos": "sustentacao do vinculo organizacional",
    "retencao de talentos": "sustentacao do vinculo organizacional",
    "retenção": "sustentacao do vinculo organizacional",
    "retencao": "sustentacao do vinculo organizacional",
    "carga de trabalho": "condicoes de trabalho a investigar",
    "turnover": "risco organizacional nao medido nesta base",
    "rotatividade": "risco organizacional nao medido nesta base",
    "reter talentos": "sustentar o vinculo organizacional",
    "intervenção": "atenção executiva",
    "intervencao": "atencao executiva",
    "intervenções": "atenções executivas",
    "intervencoes": "atencoes executivas",
    "workshop": "conversa estruturada",
    "team building": "pratica de integracao a validar",
    "feedback regular": "ritual de reconhecimento a validar",
    "sistema de feedback": "ritual de reconhecimento a validar",
    "taxa de resposta": "percentual da amostra",
}

TERMOS_QUE_INVALIDAM_HIPOTESE = {
    "condicoes de trabalho a investigar",
    "risco organizacional nao medido nesta base",
    "carga de trabalho",
    "turnover",
    "rotatividade",
    "retenção",
    "retencao",
    "reter talentos",
}


def _texto_recursivo(valor):
    if isinstance(valor, dict):
        return " ".join(_texto_recursivo(v) for v in valor.values())
    if isinstance(valor, list):
        return " ".join(_texto_recursivo(v) for v in valor)
    if valor is None:
        return ""
    return str(valor)


def _sanitizar_texto_executivo(texto, alertas):
    if not isinstance(texto, str):
        return texto
    ajustado = texto
    termos_ordenados = sorted(
        TERMOS_EXECUTIVOS_NAO_MEDIDOS.items(),
        key=lambda item: len(item[0]),
        reverse=True,
    )
    for proibido, substituto in termos_ordenados:
        if proibido.lower() in ajustado.lower():
            alertas.add(f"Termo sem medicao explicita ajustado: {proibido}")
            ajustado = re.sub(re.escape(proibido), substituto, ajustado, flags=re.IGNORECASE)
    return ajustado


def _sanitizar_json_executivo(valor, alertas):
    if isinstance(valor, dict):
        return {k: _sanitizar_json_executivo(v, alertas) for k, v in valor.items()}
    if isinstance(valor, list):
        return [_sanitizar_json_executivo(v, alertas) for v in valor]
    return _sanitizar_texto_executivo(valor, alertas)


def _indice_empresas_analiticas(pacote):
    empresas = (
        ((pacote or {}).get("analise_profunda") or {})
        .get("comparativo_empresas_mesma_holding")
        or []
    )
    indice = {}
    for item in empresas:
        if not isinstance(item, dict):
            continue
        nome = str(item.get("valor") or item.get("empresa") or "").strip().lower()
        if nome:
            indice[nome] = item
    return indice


def _normalizar_texto_validacao(valor):
    texto = str(valor or "").strip().lower()
    texto = texto.replace("'", "").replace('"', "")
    texto = texto.replace("ç", "c").replace("ã", "a").replace("á", "a").replace("à", "a")
    texto = texto.replace("â", "a").replace("é", "e").replace("ê", "e").replace("í", "i")
    texto = texto.replace("ó", "o").replace("ô", "o").replace("õ", "o").replace("ú", "u")
    return texto


def _indice_interseccoes_analiticas(pacote):
    interseccoes = (
        ((pacote or {}).get("analise_profunda") or {})
        .get("microambiente_por_interseccao_prioritaria")
        or []
    )
    linhas = []
    for item in interseccoes:
        if not isinstance(item, dict):
            continue
        rotulo = _normalizar_texto_validacao(
            item.get("rotulo") or " ".join(str(v) for v in (item.get("valores") or []))
        )
        valores = [
            _normalizar_texto_validacao(v)
            for v in (item.get("valores") or [])
            if str(v or "").strip()
        ]
        if rotulo or valores:
            linhas.append({
                "rotulo": rotulo,
                "valores": valores,
                "n": item.get("n"),
                "gap_medio": item.get("gap_medio"),
            })
    return linhas


def _filtrar_hipoteses_sem_base(revisada, alertas):
    hipoteses = revisada.get("hipoteses_e_sugestoes")
    if not isinstance(hipoteses, list):
        return
    filtradas = []
    for item in hipoteses:
        texto = _normalizar_texto_validacao(_texto_recursivo(item))
        if any(termo in texto for termo in TERMOS_QUE_INVALIDAM_HIPOTESE):
            alertas.add("Hipotese removida por depender de variavel nao medida no pacote.")
            continue
        filtradas.append(item)
    revisada["hipoteses_e_sugestoes"] = filtradas


def revisar_devolutiva_organizacional_ia(devolutiva, pacote):
    if not isinstance(devolutiva, dict):
        return devolutiva

    alertas = set()
    revisada = _sanitizar_json_executivo(devolutiva, alertas)
    empresas = _indice_empresas_analiticas(pacote)
    interseccoes = _indice_interseccoes_analiticas(pacote)
    amostra = (pacote or {}).get("amostra") or {}
    total_respondentes = amostra.get("respondentes")

    leituras = revisada.get("leitura_por_recortes")
    if isinstance(leituras, list):
        for leitura in leituras:
            if not isinstance(leitura, dict):
                continue
            recorte = str(leitura.get("recorte") or "").strip().lower()
            match = re.search(r"empresa\s*[:=]\s*([^;,+]+)", recorte)
            if not match:
                continue
            empresa = match.group(1).strip().lower()
            base = empresas.get(empresa)
            if not base:
                continue
            n_base = base.get("n")
            gap_base = base.get("gap_medio")
            comparacao = base.get("comparacao_contexto") or {}
            delta_base = comparacao.get("delta_gap_medio_vs_contexto")
            if n_base is not None and leitura.get("n") != n_base:
                leitura["n"] = n_base
                alertas.add(f"N de {empresa} reaplicado a partir da base analitica.")
            if gap_base is not None and leitura.get("gap_medio") != gap_base:
                leitura["gap_medio"] = gap_base
                alertas.add(f"Gap medio de {empresa} reaplicado a partir da base analitica.")
            if delta_base is not None:
                leitura["vs_contexto_delta"] = delta_base

    cruzamentos = revisada.get("cruzamentos_criticos")
    if isinstance(cruzamentos, list):
        cruzamentos_filtrados = []
        for cruzamento in cruzamentos:
            if not isinstance(cruzamento, dict):
                continue
            texto = _normalizar_texto_validacao(_texto_recursivo(cruzamento))
            base_encontrada = None
            for base in interseccoes:
                valores = base.get("valores") or []
                if valores and all(valor in texto for valor in valores):
                    base_encontrada = base
                    break
            if base_encontrada:
                n_base = base_encontrada.get("n")
                if n_base is not None and cruzamento.get("n") != n_base:
                    cruzamento["n"] = n_base
                    alertas.add("N de cruzamento critico reaplicado a partir da base analitica.")
                if base_encontrada.get("gap_medio") is not None:
                    cruzamento["gap_medio"] = base_encontrada.get("gap_medio")
                cruzamentos_filtrados.append(cruzamento)
            elif "spectral_a" in texto:
                alertas.add("Cruzamento com spectral_a removido por nao bater com a base analitica priorizada.")
            else:
                cruzamentos_filtrados.append(cruzamento)
        revisada["cruzamentos_criticos"] = cruzamentos_filtrados

    participacao = revisada.get("participacao_e_aderencia")
    if isinstance(participacao, dict) and total_respondentes is not None:
        for campo in ["leitura", "observacao"]:
            texto = participacao.get(campo)
            if not isinstance(texto, str):
                continue
            ajustado = re.sub(
                r"(participa[cç][aã]o(?:\s+total)?\s+foi\s+de\s+)\d+(\s+respondentes)",
                rf"\g<1>{total_respondentes}\2",
                texto,
                flags=re.IGNORECASE,
            )
            if ajustado != texto:
                participacao[campo] = ajustado
                alertas.add("Total de respondentes na participacao reaplicado a partir da base analitica.")

    _filtrar_hipoteses_sem_base(revisada, alertas)

    texto_final = _texto_recursivo(revisada).lower()
    termos_restantes = [
        termo for termo in TERMOS_EXECUTIVOS_NAO_MEDIDOS
        if termo.lower() in texto_final
    ]
    if termos_restantes:
        alertas.add(
            "Ainda ha termos que exigem revisao humana: "
            + ", ".join(sorted(set(termos_restantes)))
        )

    if alertas:
        revisada["revisao_qualidade"] = {
            "status": "ajustada_automaticamente",
            "alertas": sorted(alertas),
            "orientacao": (
                "A IA gerou trechos com inferencias ou metricas que exigiam conferencia. "
                "O bot reaplicou guardrails antes de devolver a resposta."
            ),
        }

    return revisada



def salvar_relatorio_analitico_no_supabase(dados, empresa, codrodada, email_lider, tipo):
    url = f"{SUPABASE_REST_URL}/relatorios_gerados"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "empresa": empresa,
        "codrodada": codrodada,
        "emaillider": email_lider,
        "tipo_relatorio": tipo,
        "dados_json": dados,
        "data_criacao": datetime.utcnow().isoformat()
    }
    response = requests.post(url, headers=headers, json=payload)
    response.raise_for_status()

def buscar_json_supabase(tipo_relatorio, empresa, rodada, email_lider):
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}"
    }
    url = f"{SUPABASE_REST_URL}/relatorios_gerados"
    params = {
        "empresa": f"eq.{empresa}",
        "codrodada": f"eq.{rodada}",
        "emaillider": f"eq.{email_lider}",
        "tipo_relatorio": f"eq.{tipo_relatorio}",
        "order": "data_criacao.desc",
        "limit": 1
    }
    resp = requests.get(url, headers=headers, params=params)
    if resp.status_code == 200:
        dados = resp.json()
        if dados:
            dados_json = dados[0].get("dados_json")
            if isinstance(dados_json, str):
                try:
                    dados_json = json.loads(dados_json)
                except Exception as e:
                    print("Erro ao converter dados_json:", e)
                    return None
            return dados_json
    return None

def buscar_json_microambiente(tipo_relatorio, empresa, rodada, email_lider):
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}"
    }
    url = f"{SUPABASE_REST_URL}/relatorios_gerados"
    params = {
        "empresa": f"ilike.{empresa}",
        "codrodada": f"ilike.{rodada}",
        "emaillider": f"ilike.{email_lider}",
        "tipo_relatorio": f"eq.{tipo_relatorio}",
        "order": "data_criacao.desc",
        "limit": 1
    }
    resp = requests.get(url, headers=headers, params=params)
    if resp.status_code == 200:
        dados = resp.json()
        if dados:
            dados_json = dados[0].get("dados_json")
            if isinstance(dados_json, str):
                try:
                    dados_json = json.loads(dados_json)
                except Exception as e:
                    print("Erro ao converter dados_json:", e)
                    return None
            return dados_json
    return None


def leadertrack_todos_lideres(value):
    texto = str(value or "").strip().lower()
    return texto in {
        "todos",
        "todas",
        "all",
        "__todos__",
        "__todos_lideres__",
        "__todos_lideres_contexto__",
        "__todas_empresas_contexto__",
    }


def normalizar_dados_json_relatorio(row):
    dados_json = (row or {}).get("dados_json")
    if isinstance(dados_json, str):
        try:
            return json.loads(dados_json)
        except Exception as e:
            print("Erro ao converter dados_json consolidado:", e)
            return None
    return dados_json


def buscar_relatorios_leadertrack_contexto(tipo_relatorio, empresa, rodada, contexto_ids=None, limite=500, empresas_contexto=None):
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    url = f"{SUPABASE_REST_URL}/relatorios_gerados"
    contexto_ids = contexto_ids or {}

    def executar(params):
        resp = requests.get(url, headers=headers, params=params, timeout=30)
        if resp.status_code == 200:
            return resp.json()
        print("Busca consolidada LeaderTrack ignorada:", resp.status_code, resp.text[:500])
        return None

    params_base = {
        "select": "empresa,codrodada,emaillider,tipo_relatorio,dados_json,data_criacao",
        "codrodada": f"eq.{rodada}",
        "tipo_relatorio": f"eq.{tipo_relatorio}",
        "order": "data_criacao.desc",
        "limit": str(limite),
    }
    empresas_contexto = [
        str(item or "").strip().lower()
        for item in (empresas_contexto or [])
        if str(item or "").strip()
    ]
    if empresa and not leadertrack_todos_lideres(empresa):
        params_base["empresa"] = f"eq.{empresa}"
    elif empresas_contexto:
        params_base["empresa"] = "in.(" + ",".join(empresas_contexto) + ")"

    empresa_especifica = bool(empresa and not leadertrack_todos_lideres(empresa))
    tem_contexto = any(contexto_ids.get(campo) for campo in ("filial_id", "empresa_id", "holding_id", "cliente_id"))
    tentativas = []
    for campo in ("filial_id", "empresa_id", "holding_id", "cliente_id"):
        valor = contexto_ids.get(campo)
        if valor:
            params = dict(params_base)
            params[campo] = f"eq.{valor}"
            tentativas.append(params)
    if empresa_especifica or not tem_contexto or empresas_contexto:
        tentativas.append(dict(params_base))

    rows = []
    for params in tentativas:
        dados = executar(params)
        if dados:
            rows = dados
            break

    por_lider = {}
    for row in rows or []:
        email = str(row.get("emaillider") or "").strip().lower()
        if not email or email in por_lider:
            continue
        dados_json = normalizar_dados_json_relatorio(row)
        if dados_json:
            por_lider[email] = {
                "empresa": row.get("empresa"),
                "emaillider": email,
                "dados_json": dados_json,
            }

    return list(por_lider.values())


def workbook_rows(filename):
    path = os.path.join(os.path.dirname(__file__), filename)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        return [
            {headers[index]: value for index, value in enumerate(row)}
            for row in rows
        ]
    finally:
        wb.close()


def csv_rows(filename, delimiter=";"):
    path = os.path.join(os.path.dirname(__file__), filename)
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


_MATRIZ_ARQUETIPOS_CACHE = None
_MATRIZ_MICRO_CACHE = None
_PONTOS_MAXIMOS_DIMENSAO_MICRO_CACHE = None
_PONTOS_MAXIMOS_SUBDIMENSAO_MICRO_CACHE = None
_PARES_ARQUETIPOS_CACHE = None
_SAUDE_EMOCIONAL_CACHE = None


def carregar_matriz_arquetipos_rows():
    global _MATRIZ_ARQUETIPOS_CACHE
    if _MATRIZ_ARQUETIPOS_CACHE is None:
        _MATRIZ_ARQUETIPOS_CACHE = workbook_rows("TABELA_GERAL_ARQUETIPOS_COM_CHAVE.xlsx")
    return _MATRIZ_ARQUETIPOS_CACHE


def carregar_matriz_micro_rows():
    global _MATRIZ_MICRO_CACHE
    if _MATRIZ_MICRO_CACHE is None:
        _MATRIZ_MICRO_CACHE = workbook_rows("TABELA_GERAL_MICROAMBIENTE_COM_CHAVE.xlsx")
    return _MATRIZ_MICRO_CACHE


def carregar_pontos_maximos_micro_rows(campo):
    global _PONTOS_MAXIMOS_DIMENSAO_MICRO_CACHE
    global _PONTOS_MAXIMOS_SUBDIMENSAO_MICRO_CACHE
    if campo == "DIMENSAO":
        if _PONTOS_MAXIMOS_DIMENSAO_MICRO_CACHE is None:
            _PONTOS_MAXIMOS_DIMENSAO_MICRO_CACHE = workbook_rows(
                "pontos_maximos_dimensao_microambiente.xlsx"
            )
        return _PONTOS_MAXIMOS_DIMENSAO_MICRO_CACHE
    if campo == "SUBDIMENSAO":
        if _PONTOS_MAXIMOS_SUBDIMENSAO_MICRO_CACHE is None:
            _PONTOS_MAXIMOS_SUBDIMENSAO_MICRO_CACHE = workbook_rows(
                "pontos_maximos_subdimensao_microambiente.xlsx"
            )
        return _PONTOS_MAXIMOS_SUBDIMENSAO_MICRO_CACHE
    return []


def carregar_pares_arquetipos():
    global _PARES_ARQUETIPOS_CACHE
    if _PARES_ARQUETIPOS_CACHE is None:
        path = os.path.join(os.path.dirname(__file__), "arquetipos_dominantes_por_questao.json")
        with open(path, "r", encoding="utf-8") as handle:
            _PARES_ARQUETIPOS_CACHE = json.load(handle)
    return _PARES_ARQUETIPOS_CACHE


def carregar_saude_emocional_rows():
    global _SAUDE_EMOCIONAL_CACHE
    if _SAUDE_EMOCIONAL_CACHE is None:
        _SAUDE_EMOCIONAL_CACHE = csv_rows("TABELA_SAUDE_EMOCIONAL.csv")
    return _SAUDE_EMOCIONAL_CACHE


def normalizar_recorte_valor(value):
    texto = str(value or "").strip()
    return texto if texto else "Não informado"


def valor_numerico(value, default=None):
    try:
        if value in (None, ""):
            return default
        return float(str(value).replace("%", "").replace(",", ".").strip())
    except Exception:
        return default


def arredondar_escala_likert(value, minimo=1, maximo=6):
    try:
        numero = float(value)
    except Exception:
        return None
    return int(max(minimo, min(maximo, np.floor(numero + 0.5))))


def buscar_consolidados_leadertrack_contexto(tabela, empresa, rodada, contexto_ids=None, limite=500, empresas_contexto=None):
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    url = f"{SUPABASE_REST_URL}/{tabela}"
    contexto_ids = contexto_ids or {}
    empresa_especifica = bool(empresa and not leadertrack_todos_lideres(empresa))
    empresas_contexto = [
        str(item or "").strip().lower()
        for item in (empresas_contexto or [])
        if str(item or "").strip()
    ]

    params_base = {
        "select": "*",
        "codrodada": f"eq.{rodada}",
        "order": "id.desc",
        "limit": str(limite),
    }
    if empresa_especifica:
        params_base["empresa"] = f"eq.{empresa}"
    elif empresas_contexto:
        params_base["empresa"] = "in.(" + ",".join(empresas_contexto) + ")"

    tentativas = []
    for campo in ("filial_id", "empresa_id", "holding_id", "cliente_id"):
        valor = contexto_ids.get(campo)
        if valor:
            params = dict(params_base)
            params[campo] = f"eq.{valor}"
            tentativas.append(params)
    if empresa_especifica or not tentativas or empresas_contexto:
        tentativas.append(dict(params_base))

    for params in tentativas:
        resp = requests.get(url, headers=headers, params=params, timeout=30)
        if resp.status_code == 200 and resp.json():
            return resp.json()
        if resp.status_code not in (200, 404):
            print(f"Busca {tabela} ignorada:", resp.status_code, resp.text[:500])
    return []


def respostas_arquetipos_consolidadas(rows):
    autoavaliacoes = []
    avaliacoes_equipe = []
    lideres = set()
    for row in rows or []:
        dados = normalizar_dados_json_relatorio(row) or {}
        auto = dados.get("autoavaliacao") or {}
        if isinstance(auto, dict) and isinstance(auto.get("respostas"), dict):
            autoavaliacoes.append(auto.get("respostas"))
            if auto.get("emailLider"):
                lideres.add(str(auto.get("emailLider")).strip().lower())
        for membro in dados.get("avaliacoesEquipe") or []:
            if not isinstance(membro, dict) or not isinstance(membro.get("respostas"), dict):
                continue
            avaliacoes_equipe.append(membro.get("respostas"))
            if membro.get("emailLider"):
                lideres.add(str(membro.get("emailLider")).strip().lower())
    return autoavaliacoes, avaliacoes_equipe, lideres


def calcular_arquetipos_respostas(respostas_lista):
    matriz = carregar_matriz_arquetipos_rows()
    por_chave = {str(row.get("CHAVE") or ""): row for row in matriz}
    arquetipos = ["Imperativo", "Resoluto", "Cuidativo", "Consultivo", "Prescritivo", "Formador"]
    respostas_por_questao = {}

    for respostas in respostas_lista or []:
        for questao, estrelas in (respostas or {}).items():
            questao = str(questao or "").strip()
            if not questao.startswith("Q"):
                continue
            try:
                respostas_por_questao.setdefault(questao, []).append(float(estrelas))
            except Exception:
                continue

    pontos = {name: 0.0 for name in arquetipos}
    maximos = {name: 0.0 for name in arquetipos}
    for questao, valores in respostas_por_questao.items():
        if not valores:
            continue
        estrelas_int = int(round(float(np.mean(valores))))
        estrelas_int = max(1, min(6, estrelas_int))
        for arquetipo in arquetipos:
            row = por_chave.get(f"{arquetipo}{estrelas_int}{questao}")
            if not row:
                continue
            pontos[arquetipo] += float(row.get("PONTOS_OBTIDOS") or 0)
            maximos[arquetipo] += float(row.get("PONTOS_MAXIMOS") or 0)

    return {
        arquetipo: round((pontos[arquetipo] / maximos[arquetipo]) * 100, 2)
        for arquetipo in arquetipos
        if maximos[arquetipo] > 0
    }


def calcular_arquetipos_respostas_por_respondente(respostas_lista):
    matriz = carregar_matriz_arquetipos_rows()
    por_chave = {str(row.get("CHAVE") or ""): row for row in matriz}
    arquetipos = ["Imperativo", "Resoluto", "Cuidativo", "Consultivo", "Prescritivo", "Formador"]
    acumulado = {name: [] for name in arquetipos}

    for respostas in respostas_lista or []:
        pontos = {name: 0.0 for name in arquetipos}
        maximos = {name: 0.0 for name in arquetipos}
        for questao, estrelas in (respostas or {}).items():
            questao = str(questao or "").strip()
            if not questao.startswith("Q"):
                continue
            estrelas_int = arredondar_escala_likert(estrelas)
            if estrelas_int is None:
                continue
            for arquetipo in arquetipos:
                row = por_chave.get(f"{arquetipo}{estrelas_int}{questao}")
                if not row:
                    continue
                pontos[arquetipo] += float(row.get("PONTOS_OBTIDOS") or 0)
                maximos[arquetipo] += float(row.get("PONTOS_MAXIMOS") or 0)
        for arquetipo in arquetipos:
            if maximos[arquetipo] > 0:
                acumulado[arquetipo].append((pontos[arquetipo] / maximos[arquetipo]) * 100)

    return {
        arquetipo: round(float(np.mean(valores)), 2)
        for arquetipo, valores in acumulado.items()
        if valores
    }


def classificar_percentual_arquetipo(percentual):
    try:
        valor = float(percentual)
    except Exception:
        valor = 0.0
    if valor >= 84:
        return "Muito favorável"
    if valor >= 67:
        return "Favorável"
    if valor >= 50:
        return "Pouco favorável"
    if valor >= 34:
        return "Pouco desfavorável"
    if valor >= 17:
        return "Desfavorável"
    return "Muito desfavorável"


def classificar_estrelas_arquetipo(estrelas):
    try:
        valor = float(estrelas)
    except Exception:
        valor = 0.0
    if valor >= 5.5:
        return "Muito favorável"
    if valor >= 4.5:
        return "Favorável"
    if valor >= 3.5:
        return "Pouco favorável"
    if valor >= 2.5:
        return "Pouco desfavorável"
    if valor >= 1.5:
        return "Desfavorável"
    if valor > 0:
        return "Muito desfavorável"
    return ""


def valor_coluna_matriz(row, nomes):
    chave = chave_por_nome(row, nomes)
    return row.get(chave) if chave else None


def percentual_tendencia_arquetipo(row):
    valor = valor_numerico(valor_coluna_matriz(row, ["% Tendência", "% Tendencia", "PERCETUAL"]), None)
    if valor is None:
        return None
    return round(valor * 100, 2) if valor <= 1 else round(valor, 2)


def calcular_arquetipos_analitico_respostas(autoavaliacoes, avaliacoes_equipe):
    matriz = carregar_matriz_arquetipos_rows()
    por_chave = {str(row.get("CHAVE") or ""): row for row in matriz}
    pares_por_questao = carregar_pares_arquetipos()

    def extrair_valor(questao, nota, pares):
        try:
            nota = int(round(float(nota)))
        except (TypeError, ValueError):
            return None
        if nota < 1 or nota > 6:
            return None

        # Cada afirmação pertence somente ao par oficial de arquétipos da matriz.
        for arquetipo in sorted(pares):
            row = por_chave.get(f"{arquetipo}{nota}{questao}")
            percentual = percentual_tendencia_arquetipo(row) if row else None
            if percentual is None:
                continue
            tendencia = ""
            for coluna, valor in row.items():
                nome = str(coluna or "").strip()
                normalizado = re.sub(r"[^a-zA-Z0-9]", "", nome).lower()
                if not nome.startswith("%") and normalizado in ("tendncia", "tendencia"):
                    tendencia = str(valor or "")
                    break
            return {
                "tendencia": tendencia,
                "percentual": round(float(percentual), 1),
                "afirmacao": valor_coluna_matriz(row, ["AFIRMACAO", "AFIRMAÇÃO"]) or questao,
            }
        return None

    linhas = []
    for questao, pares in pares_por_questao.items():
        grupo = " e ".join(sorted(pares))

        notas_auto = []
        for respostas in autoavaliacoes or []:
            try:
                nota = float((respostas or {}).get(questao))
            except (TypeError, ValueError):
                continue
            if 1 <= nota <= 6:
                notas_auto.append(nota)
        # Única adaptação para o consolidado: média das autoavaliações antes da matriz.
        info_auto = extrair_valor(questao, float(np.mean(notas_auto)), pares) if notas_auto else None

        percentuais_equipe = []
        soma_notas = 0
        qtd_notas = 0
        for respostas in avaliacoes_equipe or []:
            try:
                nota = int(round(float((respostas or {}).get(questao, 0))))
            except (TypeError, ValueError):
                continue
            if not 1 <= nota <= 6:
                continue
            info_individual = extrair_valor(questao, nota, pares)
            if not info_individual:
                continue
            percentuais_equipe.append(info_individual["percentual"])
            soma_notas += nota
            qtd_notas += 1

        info_equipe = None
        if percentuais_equipe:
            tendencia = extrair_valor(questao, round(soma_notas / qtd_notas), pares)
            if tendencia:
                info_equipe = {
                    "tendencia": tendencia["tendencia"],
                    "percentual": round(sum(percentuais_equipe) / len(percentuais_equipe), 2),
                    "afirmacao": tendencia["afirmacao"],
                }

        if not info_auto and not info_equipe:
            continue
        linhas.append({
            "grupoArquetipo": grupo,
            "codigo": questao,
            "afirmacao": info_auto["afirmacao"] if info_auto else info_equipe["afirmacao"],
            "autoavaliacao": {
                "tendencia": info_auto["tendencia"] if info_auto else "-",
                "percentual": info_auto["percentual"] if info_auto else 0,
            },
            "mediaEquipe": {
                "tendencia": info_equipe["tendencia"] if info_equipe else "-",
                "percentual": info_equipe["percentual"] if info_equipe else 0,
            },
        })

    return {
        "analitico": linhas,
        "escopo": "todos_lideres_contexto",
        "quantidade_autoavaliacoes": len(autoavaliacoes or []),
        "quantidade_respostas_equipe": len(avaliacoes_equipe or []),
        "fonte": "consolidado_arquetipos",
    }


def consolidar_arquetipos_respostas(rows):
    autoavaliacoes, avaliacoes_equipe, lideres = respostas_arquetipos_consolidadas(rows)
    return {
        "autoavaliacao": calcular_arquetipos_respostas(autoavaliacoes),
        "mediaEquipe": calcular_arquetipos_respostas_por_respondente(avaliacoes_equipe),
        "escopo": "todos_lideres_contexto",
        "quantidade_lideres_consolidados": len(lideres),
        "quantidade_autoavaliacoes": len(autoavaliacoes),
        "quantidade_respostas_equipe": len(avaliacoes_equipe),
        "fonte": "consolidado_arquetipos",
        "analitico": calcular_arquetipos_analitico_respostas(autoavaliacoes, avaliacoes_equipe),
    }, lideres


def respostas_micro_consolidadas(rows):
    respostas = []
    lideres = set()
    for row in rows or []:
        dados = normalizar_dados_json_relatorio(row) or {}
        for membro in dados.get("avaliacoesEquipe") or []:
            if not isinstance(membro, dict):
                continue
            if any(str(k).startswith("Q") for k in membro.keys()):
                respostas.append(membro)
            if membro.get("emailLider"):
                lideres.add(str(membro.get("emailLider")).strip().lower())
    return respostas, lideres


def respostas_micro_auto_consolidadas(rows):
    respostas = []
    lideres = set()
    for row in rows or []:
        dados = normalizar_dados_json_relatorio(row) or {}
        auto = dados.get("autoavaliacao") or {}
        if not isinstance(auto, dict):
            continue
        bloco_respostas = auto.get("respostas") if isinstance(auto.get("respostas"), dict) else auto
        if any(str(k).startswith("Q") for k in (bloco_respostas or {}).keys()):
            respostas.append(bloco_respostas)
        email = auto.get("emailLider") or row.get("emaillider")
        if email:
            lideres.add(str(email).strip().lower())
    return respostas, lideres


def registros_arquetipos_consolidados(rows):
    registros = []
    for row in rows or []:
        dados = normalizar_dados_json_relatorio(row) or {}
        auto = dados.get("autoavaliacao") or {}
        if isinstance(auto, dict) and isinstance(auto.get("respostas"), dict):
            registros.append({
                "tipo": "autoavaliacao",
                "respostas": auto.get("respostas") or {},
                "empresa": normalizar_recorte_valor(auto.get("empresa") or row.get("empresa")),
                "email_lider": str(auto.get("emailLider") or row.get("emaillider") or "").strip().lower(),
                "sexo": normalizar_recorte_valor(auto.get("sexo")),
                "etnia": normalizar_recorte_valor(auto.get("etnia")),
                "departamento": normalizar_recorte_valor(auto.get("departamento")),
                "cargo": normalizar_recorte_valor(auto.get("cargo")),
            })
        for membro in dados.get("avaliacoesEquipe") or []:
            if not isinstance(membro, dict) or not isinstance(membro.get("respostas"), dict):
                continue
            registros.append({
                "tipo": "equipe",
                "respostas": membro.get("respostas") or {},
                "empresa": normalizar_recorte_valor(membro.get("empresa") or row.get("empresa")),
                "email_lider": str(membro.get("emailLider") or row.get("emaillider") or "").strip().lower(),
                "sexo": normalizar_recorte_valor(membro.get("sexo")),
                "etnia": normalizar_recorte_valor(membro.get("etnia")),
                "departamento": normalizar_recorte_valor(membro.get("departamento")),
                "cargo": normalizar_recorte_valor(membro.get("cargo")),
            })
    return registros


def registros_micro_consolidados(rows):
    registros = []
    for row in rows or []:
        dados = normalizar_dados_json_relatorio(row) or {}
        auto = dados.get("autoavaliacao") or {}
        if isinstance(auto, dict):
            auto_respostas = auto.get("respostas") if isinstance(auto.get("respostas"), dict) else auto
            registros.append({
                "tipo": "autoavaliacao",
                "respostas": auto_respostas,
                "empresa": normalizar_recorte_valor(auto.get("empresa") or row.get("empresa")),
                "email_lider": str(auto.get("emailLider") or row.get("emaillider") or "").strip().lower(),
                "sexo": normalizar_recorte_valor(auto.get("sexo")),
                "etnia": normalizar_recorte_valor(auto.get("etnia")),
                "departamento": normalizar_recorte_valor(auto.get("departamento")),
                "cargo": normalizar_recorte_valor(auto.get("cargo")),
            })
        for membro in dados.get("avaliacoesEquipe") or []:
            if not isinstance(membro, dict):
                continue
            registros.append({
                "tipo": "equipe",
                "respostas": membro,
                "empresa": normalizar_recorte_valor(membro.get("empresa") or row.get("empresa")),
                "email_lider": str(membro.get("emailLider") or row.get("emaillider") or "").strip().lower(),
                "sexo": normalizar_recorte_valor(membro.get("sexo")),
                "etnia": normalizar_recorte_valor(membro.get("etnia")),
                "departamento": normalizar_recorte_valor(membro.get("departamento")),
                "cargo": normalizar_recorte_valor(membro.get("cargo")),
            })
    return registros


def mapa_saude_emocional():
    resultado = {}
    for row in carregar_saude_emocional_rows():
        tipo = str(row.get("TIPO") or "").strip().upper()
        codigo = str(row.get("COD_AFIRMACAO") or "").strip()
        dimensao = str(row.get("DIMENSAO_SAUDE_EMOCIONAL") or "").strip()
        if not tipo or not codigo or not dimensao:
            continue
        prefixo = "arq" if tipo.startswith("ARQ") else "micro" if tipo.startswith("MICRO") else ""
        if prefixo:
            resultado[f"{prefixo}_{codigo}"] = dimensao
    return resultado


def chave_por_nome(row, nomes):
    normalizados = {re.sub(r"[^a-zA-Z0-9]", "", str(k or "")).lower(): k for k in (row or {}).keys()}
    for nome in nomes:
        chave = normalizados.get(re.sub(r"[^a-zA-Z0-9]", "", nome).lower())
        if chave:
            return chave
    return None


def score_arquetipos_saude_emocional(registros):
    matriz = carregar_matriz_arquetipos_rows()
    por_chave = {str(row.get("CHAVE") or ""): row for row in matriz}
    arquetipos_por_questao = {}
    for row in matriz:
        codigo = str(row.get("COD_AFIRMACAO") or "").strip()
        arquetipo = str(row.get("ARQUETIPO") or "").strip()
        if codigo and arquetipo:
            arquetipos_por_questao.setdefault(codigo, set()).add(arquetipo)
    mapa = mapa_saude_emocional()
    categorias = {}
    for registro in registros or []:
        if registro.get("tipo") != "equipe":
            continue
        respostas = registro.get("respostas") or {}
        for key, categoria in mapa.items():
            if not key.startswith("arq_"):
                continue
            questao = key.replace("arq_", "")
            if questao not in respostas:
                continue
            try:
                estrelas = int(respostas.get(questao))
            except Exception:
                continue
            valores = []
            for arquetipo in sorted(arquetipos_por_questao.get(questao) or []):
                matriz_row = por_chave.get(f"{arquetipo}{estrelas}{questao}")
                if not matriz_row:
                    continue
                maximo = valor_numerico(matriz_row.get("PONTOS_MAXIMOS"), 0)
                if not maximo:
                    continue
                valores.append((valor_numerico(matriz_row.get("PONTOS_OBTIDOS"), 0) / maximo) * 100)
            if valores:
                categorias.setdefault(categoria, []).append(float(np.mean(valores)))
    return categorias


def score_micro_saude_emocional(registros):
    matriz = carregar_matriz_micro_rows()
    questoes = {}
    por_chave = {}
    for row in matriz:
        codigo = str(row.get("COD") or "").strip()
        real_key = str(row.get("name_real") or "").strip()
        ideal_key = str(row.get("name_ideal") or "").strip()
        if codigo and real_key and ideal_key:
            questoes.setdefault(codigo, row)
        chave = str(row.get("CHAVE") or "").strip()
        if chave:
            por_chave[chave] = row

    mapa = mapa_saude_emocional()
    categorias = {}
    for registro in registros or []:
        if registro.get("tipo") != "equipe":
            continue
        respostas = registro.get("respostas") or {}
        for key, categoria in mapa.items():
            if not key.startswith("micro_"):
                continue
            codigo = key.replace("micro_", "")
            questao = questoes.get(codigo)
            if not questao:
                continue
            real_key = str(questao.get("name_real") or "").strip()
            ideal_key = str(questao.get("name_ideal") or "").strip()
            if real_key not in respostas or ideal_key not in respostas:
                continue
            try:
                real = int(respostas.get(real_key))
                ideal = int(respostas.get(ideal_key))
            except Exception:
                continue
            row = por_chave.get(f"{codigo}_I{ideal}_R{real}")
            if not row:
                continue
            gap = valor_numerico(row.get("GAP"), 0)
            categorias.setdefault(categoria, []).append(max(0.0, 100.0 - abs(gap or 0)))
    return categorias


def consolidar_saude_emocional(registros_arq, registros_micro):
    categorias_valores = {}
    for source in (score_arquetipos_saude_emocional(registros_arq), score_micro_saude_emocional(registros_micro)):
        for categoria, valores in source.items():
            categorias_valores.setdefault(categoria, []).extend(valores or [])
    categorias = {
        categoria: round(float(np.mean(valores)), 1)
        for categoria, valores in categorias_valores.items()
        if valores
    }
    valores_validos = [valor for valor in categorias.values() if valor > 0]
    score = round(float(np.mean(valores_validos)), 1) if valores_validos else None
    if score is None:
        label = "Sem dados"
    elif score >= 95:
        label = "Excelente"
    elif score >= 85:
        label = "Ótimo"
    elif score >= 75:
        label = "Bom"
    elif score >= 65:
        label = "Regular"
    else:
        label = "Não adequado"
    return {
        "score_final": score,
        "label": label,
        "categorias": categorias,
        "n_respostas_arquetipos": len([r for r in registros_arq or [] if r.get("tipo") == "equipe"]),
        "n_respostas_microambiente": len([r for r in registros_micro or [] if r.get("tipo") == "equipe"]),
    }


def filtrar_registros_por_recorte(registros, recorte):
    campo, valor = recorte
    return [r for r in registros or [] if normalizar_recorte_valor(r.get(campo)).lower() == str(valor).strip().lower()]


def resumir_recorte_leadertrack(registros_arq, registros_micro):
    arq_auto = [r.get("respostas") or {} for r in registros_arq or [] if r.get("tipo") == "autoavaliacao"]
    arq_equipe = [r.get("respostas") or {} for r in registros_arq or [] if r.get("tipo") == "equipe"]
    micro_auto = []
    lideres_micro_auto = set()
    micro_equipe = []
    lideres_micro = set()
    for registro in registros_micro or []:
        if registro.get("tipo") == "autoavaliacao":
            micro_auto.append(dict(registro.get("respostas") or {}))
            if registro.get("email_lider"):
                lideres_micro_auto.add(registro.get("email_lider"))
            continue
        if registro.get("tipo") != "equipe":
            continue
        respostas = dict(registro.get("respostas") or {})
        if registro.get("email_lider"):
            respostas["emailLider"] = registro.get("email_lider")
            lideres_micro.add(registro.get("email_lider"))
        micro_equipe.append(respostas)

    micro_resumo = None
    if micro_equipe:
        micro_resumo, _ = consolidar_microambiente_respostas([
            {"dados_json": {"avaliacoesEquipe": micro_equipe}}
        ])

    micro_auto_resumo = None
    if micro_auto:
        micro_auto_resumo, _ = consolidar_microambiente_respostas_lista(
            micro_auto,
            lideres_micro_auto,
            media_antes_da_matriz=True,
        )

    return {
        "arquetipos": {
            "autoavaliacao": calcular_arquetipos_respostas(arq_auto) if arq_auto else {},
            "mediaEquipe": calcular_arquetipos_respostas_por_respondente(arq_equipe) if arq_equipe else {},
            "n_autoavaliacoes_lideres": len(arq_auto),
            "n_avaliacoes_equipe": len(arq_equipe),
        },
        "microambiente": {
            "analitico": micro_resumo,
            "media_dimensao": consolidar_microambiente_por_campo(micro_resumo, "DIMENSAO") if micro_resumo else None,
            "media_subdimensao": consolidar_microambiente_por_campo(micro_resumo, "SUBDIMENSAO") if micro_resumo else None,
            "autoavaliacao_analitico": micro_auto_resumo,
            "auto_media_dimensao": consolidar_microambiente_por_campo(micro_auto_resumo, "DIMENSAO") if micro_auto_resumo else None,
            "auto_media_subdimensao": consolidar_microambiente_por_campo(micro_auto_resumo, "SUBDIMENSAO") if micro_auto_resumo else None,
            "termometro_gaps": consolidar_microambiente_termometro(micro_resumo) if micro_resumo else None,
            "waterfall_gaps": consolidar_microambiente_waterfall(micro_resumo) if micro_resumo else None,
            "n_avaliacoes_equipe": len(micro_equipe),
            "n_lideres": len(lideres_micro),
            "n_autoavaliacoes_lideres": len(micro_auto),
        },
    }


def gerar_recortes_executivos(registros_arq, registros_micro, n_minimo=5):
    candidatos = []
    for campo in ("empresa", "sexo", "etnia"):
        valores = sorted({normalizar_recorte_valor(r.get(campo)) for r in (registros_arq or []) + (registros_micro or [])})
        for valor in valores:
            candidatos.append({
                "tipo": campo,
                "rotulo": f"{campo}: {valor}",
                "filtros": [(campo, valor)],
            })
    pares = sorted({
        (normalizar_recorte_valor(r.get("sexo")), normalizar_recorte_valor(r.get("etnia")))
        for r in (registros_arq or []) + (registros_micro or [])
    })
    for sexo, etnia in pares:
        candidatos.append({
            "tipo": "sexo_etnia",
            "rotulo": f"sexo: {sexo} + etnia: {etnia}",
            "filtros": [("sexo", sexo), ("etnia", etnia)],
        })

    geral = consolidar_saude_emocional(registros_arq, registros_micro)
    score_geral = geral.get("score_final")
    recortes = []
    for candidato in candidatos:
        arq = registros_arq
        micro = registros_micro
        for filtro in candidato["filtros"]:
            arq = filtrar_registros_por_recorte(arq, filtro)
            micro = filtrar_registros_por_recorte(micro, filtro)
        n = len([r for r in micro if r.get("tipo") == "equipe"]) or len([r for r in arq if r.get("tipo") == "equipe"])
        if n < n_minimo:
            continue
        saude = consolidar_saude_emocional(arq, micro)
        delta = None
        if score_geral is not None and saude.get("score_final") is not None:
            delta = round(float(saude.get("score_final")) - float(score_geral), 1)
        recortes.append({
            "tipo": candidato["tipo"],
            "rotulo": candidato["rotulo"],
            "n": n,
            "saude_emocional": saude,
            "delta_saude_vs_contexto": delta,
            "leadertrack": resumir_recorte_leadertrack(arq, micro),
        })
    recortes = sorted(
        recortes,
        key=lambda item: abs(float(item.get("delta_saude_vs_contexto") or 0)),
        reverse=True,
    )
    return {
        "n_minimo": n_minimo,
        "saude_emocional_geral": geral,
        "recortes": recortes[:20],
        "findings": [
            {
                "tipo": "saude_emocional",
                "recorte": item["rotulo"],
                "n": item["n"],
                "delta_pp": item.get("delta_saude_vs_contexto"),
                "leitura": "Diferença relevante de saúde emocional frente ao consolidado geral; tratar como hipótese de investigação, não como causalidade.",
            }
            for item in recortes
            if item.get("delta_saude_vs_contexto") is not None and abs(float(item.get("delta_saude_vs_contexto"))) >= 5
        ][:10],
    }


def consolidar_microambiente_respostas_lista(respostas_lista, lideres=None, media_antes_da_matriz=False):
    lideres = lideres or set()
    matriz = carregar_matriz_micro_rows()
    questoes = {}
    por_chave = {}
    for row in matriz:
        codigo = str(row.get("COD") or "").strip()
        real_key = str(row.get("name_real") or "").strip()
        ideal_key = str(row.get("name_ideal") or "").strip()
        if codigo and real_key and ideal_key:
            questoes.setdefault(codigo, row)
        chave = str(row.get("CHAVE") or "").strip()
        if chave:
            por_chave[chave] = row

    linhas = []
    for codigo, questao in questoes.items():
        real_key = str(questao.get("name_real") or "").strip()
        ideal_key = str(questao.get("name_ideal") or "").strip()
        reais = []
        ideais = []
        if media_antes_da_matriz:
            respostas_reais = []
            respostas_ideais = []
            for respostas in respostas_lista:
                if real_key not in respostas or ideal_key not in respostas:
                    continue
                try:
                    respostas_reais.append(float(respostas.get(real_key)))
                    respostas_ideais.append(float(respostas.get(ideal_key)))
                except Exception:
                    continue
            if not respostas_reais or not respostas_ideais:
                continue
            real = max(1, min(6, int(round(float(np.mean(respostas_reais))))))
            ideal = max(1, min(6, int(round(float(np.mean(respostas_ideais))))))
            row = por_chave.get(f"{codigo}_I{ideal}_R{real}")
            if not row:
                continue
            reais.append(float(row.get("PONTUACAO_REAL") or 0))
            ideais.append(float(row.get("PONTUACAO_IDEAL") or 0))
        else:
            for respostas in respostas_lista:
                if real_key not in respostas or ideal_key not in respostas:
                    continue
                real = arredondar_escala_likert(respostas.get(real_key))
                ideal = arredondar_escala_likert(respostas.get(ideal_key))
                if real is None or ideal is None:
                    continue
                row = por_chave.get(f"{codigo}_I{ideal}_R{real}")
                if not row:
                    continue
                reais.append(float(row.get("PONTUACAO_REAL") or 0))
                ideais.append(float(row.get("PONTUACAO_IDEAL") or 0))

        if not reais or not ideais:
            continue
        real_media = round(float(np.mean(reais)), 2)
        ideal_media = round(float(np.mean(ideais)), 2)
        gap = round(ideal_media - real_media, 2)
        linhas.append({
            "QUESTAO": codigo,
            "AFIRMACAO": questao.get("AFIRMACAO"),
            "DIMENSAO": questao.get("DIMENSAO"),
            "SUBDIMENSAO": questao.get("SUBDIMENSAO"),
            "PONTUACAO_REAL": real_media,
            "PONTUACAO_IDEAL": ideal_media,
            "GAP": gap,
            "N_RESPONDENTES_CONSOLIDADOS": min(len(reais), len(ideais)),
            "N_LIDERES_CONSOLIDADOS": len(lideres),
        })

    linhas = sorted(linhas, key=lambda item: abs(float(item.get("GAP") or 0)), reverse=True)
    return {
        "dados": linhas,
        "escopo": "todos_lideres_contexto",
        "quantidade_lideres_consolidados": len(lideres),
        "quantidade_respostas_equipe": len(respostas_lista),
        "fonte": "consolidado_microambiente",
    }, lideres


def consolidar_microambiente_respostas(rows):
    respostas_lista, lideres = respostas_micro_consolidadas(rows)
    return consolidar_microambiente_respostas_lista(respostas_lista, lideres, media_antes_da_matriz=False)


def consolidar_microambiente_autoavaliacoes(rows):
    respostas_lista, lideres = respostas_micro_auto_consolidadas(rows)
    return consolidar_microambiente_respostas_lista(respostas_lista, lideres, media_antes_da_matriz=True)


def consolidar_microambiente_por_campo(dados_microambiente, campo):
    grupos = {}
    for row in (dados_microambiente or {}).get("dados") or []:
        chave = str(row.get(campo) or "").strip()
        if not chave:
            continue
        grupo = grupos.setdefault(chave, {
            campo: chave,
            "PONTUACAO_REAL": [],
            "PONTUACAO_IDEAL": [],
        })
        for key in ("PONTUACAO_REAL", "PONTUACAO_IDEAL"):
            try:
                grupo[key].append(float(row.get(key) or 0))
            except Exception:
                continue

    pontos_maximos = {}
    for row in carregar_pontos_maximos_micro_rows(campo):
        if campo == "DIMENSAO":
            chave = str(row.get("DIMENSAO") or "").strip()
            maximo = valor_numerico(row.get("PONTOS_MAXIMOS_DIMENSAO"), 0)
        else:
            chave = str(row.get("SUBDIMENSAO") or "").strip()
            maximo = valor_numerico(row.get("PONTOS_MAXIMOS_SUBDIMENSAO"), 0)
        if chave and maximo:
            pontos_maximos[chave] = float(maximo)

    dados = []
    for chave, grupo in grupos.items():
        real = round(float(sum(grupo["PONTUACAO_REAL"])), 2)
        ideal = round(float(sum(grupo["PONTUACAO_IDEAL"])), 2)
        maximo = pontos_maximos.get(chave, 0)
        real_percentual = round((real / maximo) * 100, 1) if maximo else 0
        ideal_percentual = round((ideal / maximo) * 100, 1) if maximo else 0
        dados.append({
            campo: chave,
            "PONTUACAO_REAL": real,
            "PONTUACAO_IDEAL": ideal,
            "PONTOS_MAXIMOS": maximo,
            "REAL_%": real_percentual,
            "IDEAL_%": ideal_percentual,
            "GAP": round(ideal_percentual - real_percentual, 1),
        })

    return {
        "dados": sorted(dados, key=lambda item: abs(float(item["GAP"])), reverse=True),
        "escopo": "todos_lideres_contexto",
        "fonte": "consolidado_microambiente",
    }


def consolidar_microambiente_termometro(dados_microambiente):
    linhas = (dados_microambiente or {}).get("dados") or []
    total = len(linhas)
    qtd_gaps = 0
    for row in linhas:
        try:
            if abs(float(row.get("GAP") or 0)) >= 20:
                qtd_gaps += 1
        except Exception:
            continue
    percentual = round((qtd_gaps / total) * 100, 2) if total else 0
    if qtd_gaps <= 3:
        classificacao = "ALTO ESTÍMULO"
    elif qtd_gaps <= 6:
        classificacao = "ESTÍMULO"
    elif qtd_gaps <= 9:
        classificacao = "NEUTRO"
    elif qtd_gaps <= 12:
        classificacao = "BAIXO ESTÍMULO"
    else:
        classificacao = "DESMOTIVAÇÃO"
    return {
        "qtdGapsAcima20": qtd_gaps,
        "porcentagemGaps": percentual,
        "classificacao": classificacao,
        "total_afirmacoes": total,
        "escopo": "todos_lideres_contexto",
        "fonte": "consolidado_microambiente",
    }


def consolidar_microambiente_waterfall(dados_microambiente):
    return {
        "dados": {
            "dimensao": consolidar_microambiente_por_campo(dados_microambiente, "DIMENSAO").get("dados") or [],
            "subdimensao": consolidar_microambiente_por_campo(dados_microambiente, "SUBDIMENSAO").get("dados") or [],
        },
        "escopo": "todos_lideres_contexto",
        "fonte": "consolidado_microambiente",
    }


def media_dicts_relatorios(relatorios, chave):
    valores = {}
    for relatorio in relatorios or []:
        dados = relatorio.get("dados_json") or {}
        bloco = dados.get(chave) or {}
        if not isinstance(bloco, dict):
            continue
        for nome, valor in bloco.items():
            try:
                valores.setdefault(nome, []).append(float(valor or 0))
            except Exception:
                continue
    return {
        nome: round(float(np.mean(lista)), 2)
        for nome, lista in valores.items()
        if lista
    }


def consolidar_arquetipos_comparativo(relatorios):
    return {
        "autoavaliacao": media_dicts_relatorios(relatorios, "autoavaliacao"),
        "mediaEquipe": media_dicts_relatorios(relatorios, "mediaEquipe"),
        "escopo": "todos_lideres_contexto",
        "quantidade_lideres_consolidados": len(relatorios or []),
    }


def consolidar_microambiente_analitico(relatorios):
    grupos = {}
    for relatorio in relatorios or []:
        dados = relatorio.get("dados_json") or {}
        for row in dados.get("dados") or []:
            if not isinstance(row, dict):
                continue
            chave = (
                str(row.get("QUESTAO") or row.get("COD") or "").strip(),
                str(row.get("AFIRMACAO") or "").strip(),
                str(row.get("DIMENSAO") or "").strip(),
                str(row.get("SUBDIMENSAO") or "").strip(),
            )
            if not any(chave):
                continue
            grupo = grupos.setdefault(chave, {
                "QUESTAO": chave[0],
                "AFIRMACAO": chave[1],
                "DIMENSAO": chave[2],
                "SUBDIMENSAO": chave[3],
                "PONTUACAO_REAL": [],
                "PONTUACAO_IDEAL": [],
                "GAP": [],
            })
            for campo in ("PONTUACAO_REAL", "PONTUACAO_IDEAL", "GAP"):
                valor = row.get(campo)
                if valor in (None, ""):
                    continue
                try:
                    grupo[campo].append(float(str(valor).replace("%", "").replace(",", ".").strip()))
                except Exception:
                    continue

    linhas = []
    for grupo in grupos.values():
        real = round(float(np.mean(grupo["PONTUACAO_REAL"])), 2) if grupo["PONTUACAO_REAL"] else 0
        ideal = round(float(np.mean(grupo["PONTUACAO_IDEAL"])), 2) if grupo["PONTUACAO_IDEAL"] else 0
        gap = (
            round(float(np.mean(grupo["GAP"])), 2)
            if grupo["GAP"]
            else round(float(ideal) - float(real), 2)
        )
        linhas.append({
            "QUESTAO": grupo["QUESTAO"],
            "AFIRMACAO": grupo["AFIRMACAO"],
            "DIMENSAO": grupo["DIMENSAO"],
            "SUBDIMENSAO": grupo["SUBDIMENSAO"],
            "PONTUACAO_REAL": real,
            "PONTUACAO_IDEAL": ideal,
            "GAP": gap,
            "N_LIDERES_CONSOLIDADOS": len(relatorios or []),
        })

    linhas = sorted(linhas, key=lambda item: abs(float(item.get("GAP") or 0)), reverse=True)
    return {
        "dados": linhas,
        "escopo": "todos_lideres_contexto",
        "quantidade_lideres_consolidados": len(relatorios or []),
    }


def buscar_inputs_devolutiva_todos_lideres(empresa, codrodada, contexto_ids, empresas_contexto=None):
    arq_consolidados = buscar_consolidados_leadertrack_contexto(
        "consolidado_arquetipos",
        empresa,
        codrodada,
        contexto_ids,
        empresas_contexto=empresas_contexto,
    )
    micro_consolidados = buscar_consolidados_leadertrack_contexto(
        "consolidado_microambiente",
        empresa,
        codrodada,
        contexto_ids,
        empresas_contexto=empresas_contexto,
    )
    arq_comparativo_relatorios = buscar_relatorios_leadertrack_contexto(
        "arquetipos_grafico_comparativo",
        empresa,
        codrodada,
        contexto_ids,
        empresas_contexto=empresas_contexto,
    )
    micro_analitico_relatorios = buscar_relatorios_leadertrack_contexto(
        "microambiente_analitico",
        empresa,
        codrodada,
        contexto_ids,
        empresas_contexto=empresas_contexto,
    )

    arq_consolidado = None
    micro_consolidado = None
    micro_auto_consolidado = None
    lideres_arq = set()
    lideres_micro = set()
    if arq_consolidados:
        arq_consolidado, lideres_arq = consolidar_arquetipos_respostas(arq_consolidados)
    if micro_consolidados:
        micro_consolidado, lideres_micro = consolidar_microambiente_respostas(micro_consolidados)
        micro_auto_consolidado, _ = consolidar_microambiente_autoavaliacoes(micro_consolidados)
    registros_arq = registros_arquetipos_consolidados(arq_consolidados)
    registros_micro = registros_micro_consolidados(micro_consolidados)

    dados_arquetipos_comparativo = arq_consolidado or consolidar_arquetipos_comparativo(arq_comparativo_relatorios)
    dados_arquetipos_analitico = (arq_consolidado or {}).get("analitico")
    dados_microambiente_analitico = micro_consolidado or consolidar_microambiente_analitico(micro_analitico_relatorios)

    return {
        "dados_arquetipos_comparativo": dados_arquetipos_comparativo,
        "dados_arquetipos_analitico": dados_arquetipos_analitico,
        "guia_arquetipos": None,
        "dados_microambiente_analitico": dados_microambiente_analitico,
        "dados_microambiente_auto_dimensao": consolidar_microambiente_por_campo(micro_auto_consolidado, "DIMENSAO"),
        "dados_microambiente_auto_subdimensao": consolidar_microambiente_por_campo(micro_auto_consolidado, "SUBDIMENSAO"),
        "dados_microambiente_media_dimensao": consolidar_microambiente_por_campo(dados_microambiente_analitico, "DIMENSAO"),
        "dados_microambiente_subdimensao": consolidar_microambiente_por_campo(dados_microambiente_analitico, "SUBDIMENSAO"),
        "dados_microambiente_termometro_gaps": consolidar_microambiente_termometro(dados_microambiente_analitico),
        "dados_microambiente_waterfall_gaps": consolidar_microambiente_waterfall(dados_microambiente_analitico),
        "guia_microambiente": None,
        "devolutiva_executiva": gerar_recortes_executivos(registros_arq, registros_micro),
        "metadados": {
            "escopo": "todos_lideres_contexto",
            "fonte_arquetipos": "consolidado_arquetipos" if arq_consolidado else "relatorios_gerados",
            "fonte_microambiente": "consolidado_microambiente" if micro_consolidado else "relatorios_gerados",
            "lideres_com_arquetipos": len(lideres_arq) if arq_consolidado else len(arq_comparativo_relatorios or []),
            "lideres_com_microambiente": len(lideres_micro) if micro_consolidado else len(micro_analitico_relatorios or []),
            "autoavaliacoes_arquetipos": (arq_consolidado or {}).get("quantidade_autoavaliacoes"),
            "respostas_arquetipos_equipe": (arq_consolidado or {}).get("quantidade_respostas_equipe"),
            "respostas_microambiente_equipe": (micro_consolidado or {}).get("quantidade_respostas_equipe"),
        },
    }

def guia_caderno_payload(tipo, parecer, graficos=None, metadados=None):
    graficos = graficos or {}
    html = ""
    texto = ""
    if isinstance(parecer, dict):
        html = parecer.get("conteudo_html") or parecer.get("html") or parecer.get("conteudo") or ""
        texto = parecer.get("parecer") or parecer.get("texto") or parecer.get("markdown") or ""
    elif isinstance(parecer, str):
        texto = parecer

    return {
        "tipo": tipo,
        "disponivel": bool(parecer or any(graficos.values())),
        "conteudo_html": html,
        "texto": texto,
        "dados": parecer if isinstance(parecer, dict) else None,
        "graficos": graficos,
        "metadados": metadados or {},
    }


def supabase_headers(prefer_return=True, use_service_role=False):
    key = SUPABASE_SERVICE_ROLE_KEY if use_service_role and SUPABASE_SERVICE_ROLE_KEY else SUPABASE_KEY
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if prefer_return:
        headers["Prefer"] = "return=representation"
    return headers


def supabase_insert(table, payload):
    if not SUPABASE_REST_URL or not (SUPABASE_SERVICE_ROLE_KEY or SUPABASE_KEY):
        raise RuntimeError("Supabase nao configurado no ambiente.")
    url = f"{SUPABASE_REST_URL}/{table}"
    response = requests.post(url, headers=supabase_headers(use_service_role=True), json=payload, timeout=60)
    if response.status_code >= 300:
        raise RuntimeError(f"Erro ao salvar em {table}: HTTP {response.status_code} - {response.text}")
    data = response.json()
    if isinstance(data, list) and data:
        return data[0]
    return data


def supabase_upsert(table, payload, conflict_column):
    if not SUPABASE_REST_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise RuntimeError("Supabase service role nao configurado no ambiente.")
    url = f"{SUPABASE_REST_URL}/{table}"
    headers = supabase_headers(use_service_role=True)
    headers["Prefer"] = "resolution=merge-duplicates,return=representation"
    response = requests.post(
        url,
        headers=headers,
        params={"on_conflict": conflict_column},
        json=payload,
        timeout=90,
    )
    if response.status_code >= 300:
        raise RuntimeError(f"Erro ao atualizar {table}: HTTP {response.status_code} - {response.text}")
    data = response.json()
    if isinstance(data, list) and data:
        return data[0]
    return data


def supabase_patch(table, row_id, payload):
    if not SUPABASE_REST_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise RuntimeError("Supabase service role nao configurado no ambiente.")
    url = f"{SUPABASE_REST_URL}/{table}"
    response = requests.patch(
        url,
        headers=supabase_headers(use_service_role=True),
        params={"id": f"eq.{row_id}"},
        json=payload,
        timeout=60,
    )
    if response.status_code >= 300:
        raise RuntimeError(f"Erro ao atualizar {table}: HTTP {response.status_code} - {response.text}")
    data = response.json()
    if isinstance(data, list) and data:
        return data[0]
    return data


def listar_lideres_relatorios(empresa, codrodada):
    if not SUPABASE_REST_URL or not SUPABASE_KEY:
        raise RuntimeError("Supabase nao configurado no ambiente.")

    url = f"{SUPABASE_REST_URL}/relatorios_gerados"
    params = {
        "select": "emaillider,tipo_relatorio,data_criacao,dados_json",
        "empresa": f"ilike.{empresa}",
        "codrodada": f"ilike.{codrodada}",
        "tipo_relatorio": "in.(microambiente_analitico,arquetipos_analitico)",
        "order": "emaillider.asc,data_criacao.desc",
        "limit": 1000,
    }
    response = requests.get(url, headers=supabase_headers(prefer_return=False, use_service_role=True), params=params, timeout=60)
    if response.status_code >= 300:
        raise RuntimeError(f"Erro ao listar lideres: HTTP {response.status_code} - {response.text}")

    leaders = {}
    for row in response.json() or []:
        email = str(row.get("emaillider") or "").strip().lower()
        if not email:
            continue
        current = leaders.setdefault(email, {
            "email": email,
            "nome": None,
            "rotulo": email,
            "relatorios_disponiveis": [],
            "ultima_atualizacao": None,
        })
        tipo = row.get("tipo_relatorio")
        if tipo and tipo not in current["relatorios_disponiveis"]:
            current["relatorios_disponiveis"].append(tipo)
        data_criacao = row.get("data_criacao")
        if data_criacao and (not current["ultima_atualizacao"] or data_criacao > current["ultima_atualizacao"]):
            current["ultima_atualizacao"] = data_criacao

    lista = []
    for leader in leaders.values():
        if leader["nome"]:
            leader["rotulo"] = f"{leader['nome']} - {leader['email']}"
        lista.append(leader)
    return sorted(lista, key=lambda item: item["rotulo"].lower())


def listar_empresas_relatorios(codrodada=None):
    if not SUPABASE_REST_URL or not SUPABASE_KEY:
        raise RuntimeError("Supabase nao configurado no ambiente.")

    url = f"{SUPABASE_REST_URL}/relatorios_gerados"
    params = {
        "select": "empresa,codrodada,data_criacao",
        "tipo_relatorio": "in.(microambiente_analitico,arquetipos_analitico)",
        "order": "empresa.asc,data_criacao.desc",
        "limit": 1000,
    }
    if codrodada:
        params["codrodada"] = f"ilike.{codrodada}"

    response = requests.get(url, headers=supabase_headers(prefer_return=False), params=params, timeout=60)
    if response.status_code >= 300:
        raise RuntimeError(f"Erro ao listar empresas: HTTP {response.status_code} - {response.text}")

    empresas = {}
    for row in response.json() or []:
        codigo = str(row.get("empresa") or "").strip()
        if not codigo:
            continue
        key = codigo.lower()
        current = empresas.setdefault(key, {
            "codigo": key,
            "empresa": key,
            "nome": codigo.upper(),
            "rotulo": codigo.upper(),
            "rodadas": set(),
            "ultima_atualizacao": None,
        })
        rodada = row.get("codrodada")
        if rodada:
            current["rodadas"].add(str(rodada))
        data_criacao = row.get("data_criacao")
        if data_criacao and (not current["ultima_atualizacao"] or data_criacao > current["ultima_atualizacao"]):
            current["ultima_atualizacao"] = data_criacao

    lista = []
    for empresa in empresas.values():
        empresa["rodadas"] = sorted(empresa["rodadas"])
        lista.append(empresa)
    return sorted(lista, key=lambda item: item["rotulo"].lower())


def listar_rodadas_relatorios(empresa=None, email_lider=None, contexto_ids=None):
    if not SUPABASE_REST_URL or not SUPABASE_KEY:
        raise RuntimeError("Supabase nao configurado no ambiente.")

    contexto_ids = contexto_ids or {}
    rodadas = {}

    def add_rodada(row, fonte):
        codigo = str(row.get("codrodada") or "").strip()
        if not codigo:
            return
        key = codigo.lower()
        current = rodadas.setdefault(key, {
            "codigo": key,
            "codrodada": key,
            "rotulo": codigo,
            "empresas": set(),
            "lideres": set(),
            "relatorios_disponiveis": set(),
            "fontes": set(),
            "ultima_atualizacao": None,
        })
        empresa_row = row.get("empresa")
        if empresa_row:
            current["empresas"].add(str(empresa_row).strip().lower())
        lider_row = row.get("emaillider") or row.get("email")
        if lider_row:
            current["lideres"].add(str(lider_row).strip().lower())
        tipo = row.get("tipo_relatorio") or row.get("tipo")
        if tipo:
            current["relatorios_disponiveis"].add(str(tipo))
        current["fontes"].add(fonte)
        data_criacao = row.get("data_criacao") or row.get("data")
        if data_criacao and (not current["ultima_atualizacao"] or data_criacao > current["ultima_atualizacao"]):
            current["ultima_atualizacao"] = data_criacao

    limit = 1000
    offset = 0
    while True:
        url = f"{SUPABASE_REST_URL}/relatorios_gerados"
        params = {
            "select": "empresa,codrodada,emaillider,tipo_relatorio,data_criacao",
            "tipo_relatorio": "in.(microambiente_analitico,arquetipos_analitico)",
            "order": "data_criacao.desc",
            "limit": limit,
            "offset": offset,
        }
        if empresa:
            params["empresa"] = f"ilike.{empresa}"
        if email_lider:
            params["emaillider"] = f"ilike.{email_lider}"

        response = requests.get(url, headers=supabase_headers(prefer_return=False, use_service_role=True), params=params, timeout=60)
        if response.status_code >= 300:
            raise RuntimeError(f"Erro ao listar rodadas: HTTP {response.status_code} - {response.text}")

        rows = response.json() or []
        for row in rows:
            add_rodada(row, "relatorios_gerados")

        if len(rows) < limit:
            break
        offset += limit

    for table in ("relatorios_microambiente", "relatorios_arquetipos"):
        limit = 1000
        offset = 0
        while True:
            url = f"{SUPABASE_REST_URL}/{table}"
            params = {
                "select": "codrodada,email,tipo,data,nome",
                "order": "data.desc",
                "limit": limit,
                "offset": offset,
            }
            if email_lider:
                params["email"] = f"ilike.{email_lider}"
            response = requests.get(url, headers=supabase_headers(prefer_return=False, use_service_role=True), params=params, timeout=60)
            if response.status_code == 404:
                break
            if response.status_code >= 300:
                print(f"Fonte de rodadas ignorada ({table}):", response.status_code, response.text)
                break

            rows = response.json() or []
            for row in rows:
                add_rodada(row, table)
            if len(rows) < limit:
                break
            offset += limit

    url = f"{SUPABASE_REST_URL}/leadertrack_rodadas"
    params = {
        "select": "codrodada,cliente_id,holding_id,ciclo_avaliacao_id",
        "order": "codrodada.desc",
        "limit": 1000,
    }
    if contexto_ids.get("cliente_id"):
        params["cliente_id"] = f"eq.{contexto_ids.get('cliente_id')}"
    if contexto_ids.get("holding_id"):
        params["holding_id"] = f"eq.{contexto_ids.get('holding_id')}"
    if contexto_ids.get("ciclo_avaliacao_id"):
        params["ciclo_avaliacao_id"] = f"eq.{contexto_ids.get('ciclo_avaliacao_id')}"
    response = requests.get(url, headers=supabase_headers(prefer_return=False, use_service_role=True), params=params, timeout=60)
    if response.status_code < 300:
        for row in response.json() or []:
            add_rodada(row, "leadertrack_rodadas")
    elif response.status_code != 404:
        print("Fonte de rodadas ignorada (leadertrack_rodadas):", response.status_code, response.text)

    lista = []
    for rodada in rodadas.values():
        rodada["empresas"] = sorted(rodada["empresas"])
        rodada["lideres"] = sorted(rodada["lideres"])
        rodada["relatorios_disponiveis"] = sorted(rodada["relatorios_disponiveis"])
        rodada["fontes"] = sorted(rodada["fontes"])
        lista.append(rodada)
    return sorted(lista, key=lambda item: item.get("ultima_atualizacao") or "", reverse=True)


def leadertrack_cache_key(empresa, codrodada, email_lider, equipe_tipo, gap_id, etapa, semana_inicio=None, semana_fim=None):
    parts = [
        "leadertrack_pdi_v11",
        str(empresa or "").strip().lower(),
        str(codrodada or "").strip().lower(),
        str(email_lider or "").strip().lower(),
        str(equipe_tipo or "direta").strip().lower(),
        str(gap_id or "").strip().lower(),
        str(etapa or "").strip().lower(),
        str(semana_inicio or ""),
        str(semana_fim or ""),
    ]
    return "|".join(parts)


def normalizar_plano_semanal_leadertrack(resultado):
    if not isinstance(resultado, dict):
        return resultado
    semanas = resultado.get("plano_12_semanas")
    if isinstance(semanas, list) and semanas:
        return resultado

    for key in ("plano", "semana", "plano_semanal"):
        value = resultado.get(key)
        if isinstance(value, dict) and value.get("semana"):
            resultado["plano_12_semanas"] = [value]
            return resultado

    for key in ("semanas", "weeks"):
        value = resultado.get(key)
        if isinstance(value, list) and value:
            resultado["plano_12_semanas"] = value
            return resultado

    if resultado.get("semana"):
        resultado["plano_12_semanas"] = [resultado.copy()]
    return resultado


def _lista_preenchida(valor):
    return isinstance(valor, list) and any(str(item or "").strip() for item in valor)


def _roteiro_tempo_valido(valor):
    if not isinstance(valor, list) or not valor:
        return False
    for item in valor:
        if not isinstance(item, dict):
            return False
        if not str(item.get("atividade") or "").strip():
            return False
        try:
            if int(item.get("minutos") or 0) <= 0:
                return False
        except Exception:
            return False
    return True


def problemas_qualidade_plano_leadertrack(resultado):
    problemas = []
    if not isinstance(resultado, dict):
        return ["Resposta nao retornou objeto JSON."]

    semanas = resultado.get("plano_12_semanas")
    if not isinstance(semanas, list) or not semanas:
        return ["Resposta nao trouxe plano_12_semanas com semanas validas."]

    tipos_por_semana = []
    objetivos_por_semana = []
    for semana in semanas:
        if not isinstance(semana, dict):
            problemas.append("Uma semana veio em formato invalido.")
            continue

        numero = semana.get("semana") or "?"
        tipo = str(semana.get("tipo_de_intervencao") or "").strip()
        objetivo = str(semana.get("objetivo") or "").strip().lower()
        tipos_por_semana.append(tipo)
        objetivos_por_semana.append(objetivo)

        campos_texto = [
            "etapa_do_ciclo",
            "foco_da_semana",
            "assunto_especifico_das_afirmacoes",
            "speech_sugerido_do_lider",
            "diferenca_objetiva_da_semana_anterior",
            "proxima_pratica_observavel",
            "feedback_necessario_para_liberar_proxima_semana",
            "criterio_de_conclusao_da_semana",
            "arquetipo_dominante_a_acionar",
            "como_usar_arquetipo_dominante",
            "arquetipo_complementar_a_desenvolver",
            "pratica_para_desenvolver_arquetipo",
            "indicador",
            "resultado_esperado",
        ]
        for campo in campos_texto:
            if not str(semana.get(campo) or "").strip():
                problemas.append(f"Semana {numero}: campo ausente ou vazio: {campo}.")

        for campo in (
            "palavras_chave_das_afirmacoes",
            "frases_especificas_para_usar",
            "comentarios_de_observacao",
            "acoes_praticas",
            "perguntas_para_equipe",
            "tarefa_do_lider",
            "tarefa_da_equipe",
            "afirmacoes_impactadas",
        ):
            if not _lista_preenchida(semana.get(campo)):
                problemas.append(f"Semana {numero}: lista ausente ou vazia: {campo}.")

        for campo, subcampos in {
            "registro_do_lider_antes_de_avancar": ("o_que_fiz", "o_que_observei", "o_que_mudou", "onde_travou"),
            "autodesenvolvimento_do_lider": ("pratica_emocional", "treino_de_speech", "leitura_curta_recomendada", "reflexao_individual"),
            "alinhamento_com_superior": ("quando_fazer", "objetivo_do_alinhamento"),
            "compromisso_de_agenda": ("titulo", "tipo", "duracao_minutos", "descricao"),
        }.items():
            bloco = semana.get(campo)
            if not isinstance(bloco, dict):
                problemas.append(f"Semana {numero}: bloco ausente ou invalido: {campo}.")
                continue
            for subcampo in subcampos:
                if not str(bloco.get(subcampo) or "").strip():
                    problemas.append(f"Semana {numero}: campo ausente ou vazio: {campo}.{subcampo}.")

        if not _roteiro_tempo_valido(semana.get("roteiro_de_tempo")):
            problemas.append(f"Semana {numero}: roteiro_de_tempo precisa ser lista de atividades com minutos.")

        try:
            total = int(semana.get("tempo_total_estimado_minutos") or 0)
            if total <= 0 or total > 120:
                problemas.append(f"Semana {numero}: tempo_total_estimado_minutos fora do limite de ate 120.")
        except Exception:
            problemas.append(f"Semana {numero}: tempo_total_estimado_minutos invalido.")

        if not tipo:
            problemas.append(f"Semana {numero}: tipo_de_intervencao vazio.")
        elif "|" in tipo or "," in tipo or "/" in tipo:
            problemas.append(f"Semana {numero}: tipo_de_intervencao deve ser uma unica opcao.")

    for index in range(1, len(tipos_por_semana)):
        if tipos_por_semana[index] and tipos_por_semana[index] == tipos_por_semana[index - 1]:
            problemas.append(f"Semanas consecutivas repetiram tipo_de_intervencao: {tipos_por_semana[index]}.")

    for index in range(1, len(objetivos_por_semana)):
        if objetivos_por_semana[index] and objetivos_por_semana[index] == objetivos_por_semana[index - 1]:
            problemas.append("Semanas consecutivas repetiram o mesmo objetivo central.")

    return problemas


def build_quality_retry_prompt(resultado, problemas):
    return (
        "Revise a resposta JSON abaixo para corrigir problemas de qualidade do PDI LeaderTrack. "
        "Mantenha o mesmo lider, tema, afirmacoes, arquetipos e semanas. "
        "Nao invente dados numericos nem exponha percentuais para a equipe. "
        "Preencha todos os campos ausentes com conteudo especifico das afirmacoes envolvidas. "
        "Inclua palavras-chave literais ou muito proximas das afirmacoes, frases prontas que o lider possa falar "
        "e comentarios de observacao com sinais concretos da rotina. "
        "Varie as intervencoes entre semanas e mantenha no maximo 120 minutos por semana. "
        "Responda somente JSON valido, preservando a estrutura do objeto original.\n\n"
        f"PROBLEMAS_ENCONTRADOS:\n{json.dumps(problemas[:30], ensure_ascii=False, indent=2)}\n\n"
        f"JSON_A_CORRIGIR:\n{json.dumps(resultado, ensure_ascii=False, indent=2)}"
    )


def revisar_plano_leadertrack_se_incompleto(resultado, prompt_base, model="gpt-4.1-mini"):
    problemas = problemas_qualidade_plano_leadertrack(resultado)
    if not problemas:
        resultado["qualidade_geracao"] = {
            "status": "aprovada",
            "problemas_corrigidos": [],
        }
        return resultado

    resultado["qualidade_geracao"] = {
        "status": "gerada_com_pendencias",
        "problemas_iniciais": problemas[:30],
        "orientacao": "Plano entregue sem segunda chamada de revisao para evitar timeout. Use regenerar se quiser tentar melhorar a qualidade.",
    }
    return resultado


def buscar_cache_leadertrack(empresa, email_lider, cache_key):
    if not SUPABASE_REST_URL or not (SUPABASE_SERVICE_ROLE_KEY or SUPABASE_KEY):
        return None

    url = f"{SUPABASE_REST_URL}/leadertrack_pdi_historico"
    params = {
        "select": "id,dados_depois,data_evento",
        "empresa": f"eq.{empresa}",
        "profissional_email": f"eq.{email_lider}",
        "origem": "eq.leadertrackbot_cache",
        "tipo_evento": "eq.ia_gerada",
        "order": "data_evento.desc",
        "limit": 80,
    }
    response = requests.get(url, headers=supabase_headers(prefer_return=False, use_service_role=True), params=params, timeout=60)
    if response.status_code >= 300:
        print("Cache LeaderTrack indisponivel:", response.status_code, response.text)
        return None

    for row in response.json() or []:
        dados = row.get("dados_depois") or {}
        if isinstance(dados, str):
            try:
                dados = json.loads(dados)
            except Exception:
                dados = {}
        if dados.get("cache_key") == cache_key:
            payload = dados.get("payload") or {}
            if isinstance(payload, dict):
                payload["persistencia"] = "cache_lido"
                payload["fonte"] = "cache"
                payload["cache"] = {
                    "status": "hit",
                    "cache_key": cache_key,
                    "historico_id": row.get("id"),
                    "data_evento": row.get("data_evento"),
                }
                return payload
    return None


def salvar_cache_leadertrack(empresa, contexto, contexto_ids, email_lider, nome_lider, cache_key, payload, gerado_por=None):
    evento = {
        "profissional_email": email_lider,
        "profissional_nome": nome_lider,
        "cliente_id": contexto_ids.get("cliente_id"),
        "holding_id": contexto_ids.get("holding_id"),
        "empresa_id": contexto_ids.get("empresa_id"),
        "filial_id": contexto_ids.get("filial_id"),
        "empresa": empresa,
        "contexto": contexto,
        "origem": "leadertrackbot_cache",
        "tipo_evento": "ia_gerada",
        "descricao_evento": "Geracao de IA LeaderTrack salva para reuso como cache tecnico.",
        "dados_antes": None,
        "dados_depois": {
            "cache_key": cache_key,
            "payload": payload,
        },
        "registrado_por": gerado_por,
    }
    try:
        saved = supabase_insert("leadertrack_pdi_historico", evento)
        return {
            "status": "salvo_no_historico_cache",
            "historico_id": saved.get("id") if isinstance(saved, dict) else None,
        }
    except Exception as exc:
        print("Erro ao salvar cache LeaderTrack:", exc)
        return {
            "status": "nao_salvo",
            "erro": str(exc),
        }


def _normalizar_chave_amostra(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def avaliar_amostra_leadertrack(*relatorios):
    textos = []
    campos = {}

    def walk(value, path=""):
        if isinstance(value, dict):
            for key, item in value.items():
                key_norm = _normalizar_chave_amostra(key)
                if isinstance(item, (int, float, str, bool)) and (
                    key_norm in (
                    "respondentes",
                    "respostasequipe",
                    "respostasdaequipe",
                    "avaliacoesequipe",
                    "avaliacoesdeequipe",
                    "avaliacoesdaequipe",
                    "elegiveismedia",
                    "elegiveisparamedia",
                    "menosde3meses",
                    "menos3meses",
                    "amostrainsuficiente",
                    )
                    or "respondente" in key_norm
                    or "respostaequipe" in key_norm
                    or "avaliacoesequipe" in key_norm
                    or "avaliacoesdeequipe" in key_norm
                    or "avaliacoesdaequipe" in key_norm
                    or "elegivel" in key_norm
                    or "elegivei" in key_norm
                    or "amostra" in key_norm
                    or "menosde3" in key_norm
                ):
                    campos[key_norm] = item
                walk(item, f"{path}.{key}" if path else str(key))
        elif isinstance(value, list):
            for index, item in enumerate(value):
                walk(item, f"{path}[{index}]")
        elif isinstance(value, str):
            textos.append(value.lower())

    for relatorio in relatorios:
        walk(relatorio)

    texto_unificado = " ".join(textos)
    def numero(value):
        try:
            return float(str(value).replace(",", "."))
        except Exception:
            return None

    elegiveis = campos.get("elegiveismedia", campos.get("elegiveisparamedia"))
    respostas = campos.get(
        "respostasequipe",
        campos.get(
            "respostasdaequipe",
            campos.get(
                "avaliacoesequipe",
                campos.get(
                    "avaliacoesdeequipe",
                    campos.get("avaliacoesdaequipe", campos.get("respondentes")),
                ),
            ),
        ),
    )
    menos_3_meses = campos.get("menosde3meses", campos.get("menos3meses"))

    for key, value in campos.items():
        if "elegivel" in key or "elegivei" in key:
            elegiveis = value if elegiveis in (None, "") else elegiveis
        if "respondente" in key or "respostaequipe" in key or "avaliacoesequipe" in key or "avaliacoesdeequipe" in key or "avaliacoesdaequipe" in key:
            respostas = value if respostas in (None, "") else respostas
        if "menosde3" in key:
            menos_3_meses = value if menos_3_meses in (None, "") else menos_3_meses

    if respostas in (None, ""):
        match = re.search(
            r"(\d+)\s*avalia[cç][oõ]es?\s*(?:da\s+|de\s+)?equipe",
            texto_unificado,
            re.IGNORECASE,
        )
        if match:
            respostas = match.group(1)

    elegiveis_num = numero(elegiveis)
    respostas_num = numero(respostas)

    insuficiente = (
        "amostra insuficiente" in texto_unificado
        or "menos de 3 respostas" in texto_unificado
        or "menos de três respostas" in texto_unificado
        or bool(campos.get("amostrainsuficiente"))
        or (elegiveis_num is not None and elegiveis_num < 3)
        or (respostas_num is not None and respostas_num < 3)
    )

    return {
        "insuficiente": bool(insuficiente),
        "respostas_equipe": respostas,
        "elegiveis_media": elegiveis,
        "menos_de_3_meses": menos_3_meses,
        "criterio": "Amostra minima recomendada: pelo menos 3 respostas elegiveis para media da equipe.",
        "orientacao": (
            "Quando a amostra e insuficiente, os graficos podem existir, mas a devolutiva deve ser tratada "
            "como leitura limitada. Evite conclusoes fortes sobre percepcao coletiva da equipe."
            if insuficiente else
            "Amostra sem sinal automatico de insuficiencia nos metadados disponiveis."
        ),
    }


def persistir_devolutiva_leadertrack(devolutiva, gerado_por=None):
    contexto_ids = devolutiva.get("contexto_ids") or {}
    parent_payload = {
        "empresa": devolutiva.get("empresa"),
        "contexto": devolutiva.get("contexto"),
        "cliente_id": contexto_ids.get("cliente_id"),
        "holding_id": contexto_ids.get("holding_id"),
        "empresa_id": contexto_ids.get("empresa_id"),
        "filial_id": contexto_ids.get("filial_id"),
        "codrodada": devolutiva.get("codrodada"),
        "email_lider": devolutiva.get("email_lider"),
        "nome_lider": devolutiva.get("nome_lider"),
        "status": "rascunho_gerado",
        "limite_gaps": len(devolutiva.get("gaps_priorizados") or []),
        "maximo_gaps_por_ciclo": (devolutiva.get("faseamento_anual_sugerido") or {}).get("maximo_recomendado_por_ciclo", 4),
        "gap_minimo_percentual": (devolutiva.get("criterios_de_leitura") or {}).get("gap_minimo_percentual"),
        "baixa_referencia_threshold_percentual": (devolutiva.get("criterios_de_leitura") or {}).get("baixa_referencia_threshold_percentual"),
        "dados_entrada_resumo": {
            "arquetipos": devolutiva.get("arquetipos"),
            "resumo_severidade": devolutiva.get("resumo_severidade"),
            "criterios_de_leitura": devolutiva.get("criterios_de_leitura"),
        },
        "todas_afirmacoes_microambiente": devolutiva.get("todas_afirmacoes_microambiente") or [],
        "baixa_referencia": devolutiva.get("baixa_referencia") or [],
        "faseamento_anual_sugerido": devolutiva.get("faseamento_anual_sugerido") or {},
        "resposta_leadertrackbot_json": devolutiva,
        "pdi_enviado_para_modulo_pdi": False,
        "meta_desempenho_sugerida": any((pdi.get("meta_desempenho_sugerida") for pdi in devolutiva.get("pdis") or [])),
        "gerado_por": gerado_por,
    }
    saved_parent = supabase_insert("leadertrack_devolutivas", parent_payload)
    devolutiva_id = saved_parent.get("id")

    for pdi in devolutiva.get("pdis") or []:
        gap = pdi.get("gap") or {}
        gap_id = pdi.get("gap_id")
        for semana in pdi.get("plano_12_semanas") or []:
            supabase_insert("leadertrack_pdi_acompanhamento", {
            "devolutiva_id": devolutiva_id,
            "cliente_id": contexto_ids.get("cliente_id"),
            "holding_id": contexto_ids.get("holding_id"),
            "empresa_id": contexto_ids.get("empresa_id"),
            "filial_id": contexto_ids.get("filial_id"),
            "gap_id": gap_id,
                "questao": gap.get("questao"),
                "dimensao": gap.get("dimensao"),
                "subdimensao": gap.get("subdimensao"),
                "afirmacao": gap.get("afirmacao"),
                "semana": semana.get("semana"),
                "foco_da_semana": semana.get("foco_da_semana"),
                "objetivo": semana.get("objetivo"),
                "prazo": semana.get("prazo"),
                "status": semana.get("status") or "nao_iniciado",
                "acoes_praticas": semana.get("acoes_praticas") or [],
                "formato_sugerido": semana.get("formato_sugerido"),
                "tarefa_do_lider": semana.get("tarefa_do_lider") or [],
                "tarefa_da_equipe": semana.get("tarefa_da_equipe") or [],
                "perguntas_para_equipe": semana.get("perguntas_para_equipe") or [],
                "o_que_mostrar_para_equipe": semana.get("o_que_mostrar_para_equipe") or [],
                "o_que_nao_mostrar_para_equipe": semana.get("o_que_nao_mostrar_para_equipe") or [],
                "indicador": semana.get("indicador"),
                "evidencia_esperada": semana.get("evidencia_esperada"),
                "resultado_esperado": semana.get("resultado_esperado"),
                "indicadores_operacionais_relacionados": semana.get("indicadores_operacionais_relacionados") or [],
                "metrica_operacional_base": semana.get("metrica_operacional_base"),
                "evolucao_operacional_observada": semana.get("evolucao_operacional_observada"),
                "resultado_obtido": semana.get("resultado_obtido"),
                "observacoes_de_evolucao": semana.get("observacoes_de_evolucao"),
            })

        historico = dict(pdi.get("historico_evento_inicial") or {})
        if historico:
            historico["devolutiva_id"] = devolutiva_id
            historico["cliente_id"] = contexto_ids.get("cliente_id")
            historico["holding_id"] = contexto_ids.get("holding_id")
            historico["empresa_id"] = contexto_ids.get("empresa_id")
            historico["filial_id"] = contexto_ids.get("filial_id")
            supabase_insert("leadertrack_pdi_historico", historico)

        meta = dict(pdi.get("meta_desempenho_sugerida") or {})
        if meta:
            meta["devolutiva_id"] = devolutiva_id
            meta["cliente_id"] = contexto_ids.get("cliente_id")
            meta["holding_id"] = contexto_ids.get("holding_id")
            meta["empresa_id"] = contexto_ids.get("empresa_id")
            meta["filial_id"] = contexto_ids.get("filial_id")
            supabase_insert("leadertrack_pdi_meta_desempenho", meta)

        indicadores = (
            (pdi.get("diagnostico") or {})
            .get("indicadores_de_efetividade", {})
            .get("indicadores_operacionais_sugeridos", [])
        )
        for indicador in indicadores:
            supabase_insert("leadertrack_pdi_indicadores_operacionais", {
                "devolutiva_id": devolutiva_id,
                "cliente_id": contexto_ids.get("cliente_id"),
                "holding_id": contexto_ids.get("holding_id"),
                "empresa_id": contexto_ids.get("empresa_id"),
                "filial_id": contexto_ids.get("filial_id"),
                "gap_id": gap_id,
                "nome_indicador": indicador,
                "fonte": "sugerido_leadertrackbot",
            })

    return {
        "devolutiva_id": devolutiva_id,
        "status_persistencia": "salvo_no_supabase",
    }

@app.route("/emitir-parecer-arquetipos", methods=["POST", "OPTIONS"])
def emitir_parecer_arquetipos():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
        return response

    try:
        dados = request.get_json()
        empresa = dados["empresa"].lower()
        rodada = dados["codrodada"].lower()
        email_lider = dados["emailLider"].lower()

        tipo_relatorio = "arquetipos_parecer_ia"

        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}"
        }

        def buscar_json(tipo):
            url = f"{SUPABASE_REST_URL}/relatorios_gerados"
            params = {
                "empresa": f"eq.{empresa}",
                "codrodada": f"eq.{rodada}",
                "emaillider": f"eq.{email_lider}",
                "tipo_relatorio": f"eq.{tipo}",
                "order": "data_criacao.desc",
                "limit": 1
            }
            resp = requests.get(url, headers=headers, params=params)
            if resp.status_code == 200 and resp.json():
                return resp.json()[0]["dados_json"]
            return None

        json_auto_vs_equipe = buscar_json("arquetipos_grafico_comparativo")
        if json_auto_vs_equipe and isinstance(json_auto_vs_equipe, str):
            try:
                json_auto_vs_equipe = json.loads(json_auto_vs_equipe)
            except:
                json_auto_vs_equipe = None

        with open("guias_completos_unificados.txt", "r", encoding="utf-8") as f:
            texto = f.read()
        inicio = texto.find("##### INICIO ARQUETIPOS #####")
        fim = texto.find("##### FIM ARQUETIPOS #####")
        guia = texto[inicio + len("##### INICIO ARQUETIPOS #####"):fim].strip() if inicio != -1 and fim != -1 else "Guia de Arquétipos não encontrado."

        marcador = "Abaixo, o resultado da análise de Arquétipos relativa ao modo como voce lidera em sua visão, comparado com a média da visão de sua equipe direta:"
        partes = guia.split(marcador)

        imagem_base64 = ""
        if json_auto_vs_equipe:
            labels = list(json_auto_vs_equipe["autoavaliacao"].keys())
            auto = list(json_auto_vs_equipe["autoavaliacao"].values())
            equipe = list(json_auto_vs_equipe["mediaEquipe"].values())
            x = np.arange(len(labels))
            fig, ax = plt.subplots(figsize=(10, 5))
            ax.bar(x - 0.2, auto, width=0.4, label="Autoavaliação", color="orange")
            ax.bar(x + 0.2, equipe, width=0.4, label="Equipe", color="lightblue")
            for i in range(len(labels)):
                ax.text(x[i] - 0.2, auto[i] + 1, f"{auto[i]:.0f}%", ha='center', fontsize=8)
                ax.text(x[i] + 0.2, equipe[i] + 1, f"{equipe[i]:.0f}%", ha='center', fontsize=8)
            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=45)
            ax.axhline(50, color="gray", linestyle="--")
            ax.axhline(60, color="gray", linestyle=":")
            ax.set_ylim(0, 100)
            ax.set_title("ARQUÉTIPOS AUTO VS EQUIPE", fontsize=14, weight="bold")
            plt.tight_layout()

            buf = io.BytesIO()
            plt.savefig(buf, format='png')
            buf.seek(0)
            imagem_base64 = base64.b64encode(buf.read()).decode("utf-8")
            plt.close()

        bloco_html = partes[0] + f"<br><br>{marcador}<br><br><img src='data:image/png;base64,{imagem_base64}' style='width:100%;max-width:800px;'><br><br>" + partes[1] if len(partes) == 2 else guia

        dados_retorno = {
            "titulo": "ARQUÉTIPOS DE GESTÃO",
            "subtitulo": f"{empresa.upper()} / {rodada.upper()} / {email_lider}",
            "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "conteudo_html": bloco_html
        }

        salvar_relatorio_analitico_no_supabase(dados_retorno, empresa, rodada, email_lider, tipo_relatorio)

        response = jsonify(dados_retorno)
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro no parecer IA arquetipos:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500

@app.route("/emitir-parecer-microambiente", methods=["POST", "OPTIONS"])
def emitir_parecer_microambiente():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
        return response

    try:
        dados = request.get_json()
        empresa = dados["empresa"].lower()
        rodada = dados["codrodada"].lower()
        email_lider = dados["emailLider"].lower()

        tipo_relatorio = "microambiente_parecer_ia"

        with open("guias_completos_unificados.txt", "r", encoding="utf-8") as f:
            texto = f.read()
        inicio = texto.find("##### INICIO MICROAMBIENTE #####")
        fim = texto.find("##### FIM MICROAMBIENTE #####")
        guia = texto[inicio + len("##### INICIO MICROAMBIENTE #####"):fim].strip() if inicio != -1 and fim != -1 else "Guia de Microambiente não encontrado."

        dados_retorno = {
            "titulo": "PARECER INTELIGENTE - MICROAMBIENTE",
            "subtitulo": f"{empresa.upper()} / {rodada.upper()} / {email_lider}",
            "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "conteudo_html": guia
        }

        salvar_relatorio_analitico_no_supabase(dados_retorno, empresa, rodada, email_lider, tipo_relatorio)

        response = jsonify(dados_retorno)
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro no parecer IA microambiente:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500

@app.route("/buscar-json-supabase", methods=["POST", "OPTIONS"])
def buscar_json_supabase_rota():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
        return response

    try:
        dados = request.get_json()
        tipo_relatorio = dados["tipo_relatorio"]
        empresa = dados["empresa"].lower()
        codrodada = dados["codrodada"].lower()
        emailLider = dados["emailLider"].lower()

        print(f"🔍 Buscando dados: {tipo_relatorio}, {empresa}, {codrodada}, {emailLider}")

        dados_json = buscar_json_supabase(tipo_relatorio, empresa, codrodada, emailLider)

        if dados_json:
            response = jsonify(dados_json)
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 200
        else:
            response = jsonify({"erro": "Dados não encontrados"})
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 404

    except Exception as e:
        print("Erro ao buscar JSON:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500

@app.route("/buscar-json-microambiente", methods=["POST", "OPTIONS"])
def buscar_json_microambiente_rota():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
        return response

    try:
        dados = request.get_json()
        tipo_relatorio = dados["tipo_relatorio"]
        empresa = dados["empresa"].lower()
        codrodada = dados["codrodada"].lower()
        emailLider = dados["emailLider"].lower()

        print(f"🔍 Buscando dados microambiente: {tipo_relatorio}, {empresa}, {codrodada}, {emailLider}")

        dados_json = buscar_json_microambiente(tipo_relatorio, empresa, codrodada, emailLider)

        if dados_json:
            response = jsonify(dados_json)
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 200
        else:
            response = jsonify({"erro": "Dados não encontrados"})
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 404

    except Exception as e:
        print("Erro ao buscar JSON microambiente:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500

@app.route("/teste-prompt-leadertrack", methods=["GET"])
def teste_prompt_leadertrack():
    try:
        prompt = carregar_prompt_leadertrack()

        return jsonify({
            "status": "ok",
            "mensagem": "Prompt Leadertrack carregado com sucesso.",
            "tamanho_caracteres": len(prompt),
            "inicio_prompt": prompt[:300]
        }), 200

    except Exception as e:
        return jsonify({
            "status": "erro",
            "mensagem": str(e)
        }), 500

@app.route("/chat-leadertrack", methods=["POST", "OPTIONS"])
def chat_leadertrack():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
        return response

    try:
        dados = request.get_json()

        empresa = dados.get("empresa", "").lower()
        codrodada = dados.get("codrodada", "").lower()
        email_lider = dados.get("emailLider", "").lower()
        pergunta = dados.get("pergunta", "")
        pagina_atual = dados.get("paginaAtual", "")
        url_atual = dados.get("urlAtual", "")

        if not empresa or not codrodada or not email_lider or not pergunta:
            response = jsonify({
                "erro": "Campos obrigatórios ausentes.",
                "campos_necessarios": [
                    "empresa",
                    "codrodada",
                    "emailLider",
                    "pergunta"
                ]
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        prompt_base = carregar_prompt_leadertrack()

        # Busca os dados reais de Arquétipos já gerados pelo Leadertrack no Supabase
        dados_arquetipos_comparativo = buscar_json_supabase(
            "arquetipos_grafico_comparativo",
            empresa,
            codrodada,
            email_lider
        )
        
        dados_arquetipos_analitico = buscar_json_supabase(
            "arquetipos_analitico",
            empresa,
            codrodada,
            email_lider
        )
        
        guia_arquetipos = buscar_json_supabase(
            "arquetipos_parecer_ia",
            empresa,
            codrodada,
            email_lider
        )
        
        # Busca os dados reais de Microambiente já gerados pelo Leadertrack no Supabase
        dados_microambiente_analitico = buscar_json_microambiente(
            "microambiente_analitico",
            empresa,
            codrodada,
            email_lider
        )
        
        dados_microambiente_subdimensao = buscar_json_microambiente(
            "microambiente_grafico_mediaequipe_subdimensao",
            empresa,
            codrodada,
            email_lider
        )
        
        dados_microambiente_termometro_gaps = buscar_json_microambiente(
            "microambiente_termometro_gaps",
            empresa,
            codrodada,
            email_lider
        )
        
        dados_microambiente_waterfall_gaps = buscar_json_microambiente(
            "microambiente_waterfall_gaps",
            empresa,
            codrodada,
            email_lider
        )
        
        guia_microambiente = buscar_json_microambiente(
            "microambiente_parecer_ia",
            empresa,
            codrodada,
            email_lider
        )
        
        resposta_ia = gerar_resposta_ia_leadertrack(
            pergunta=pergunta,
            prompt_base=prompt_base,
            empresa=empresa,
            codrodada=codrodada,
            email_lider=email_lider,
            pagina_atual=pagina_atual,
            url_atual=url_atual,            
            dados_arquetipos_comparativo=dados_arquetipos_comparativo,
            dados_arquetipos_analitico=dados_arquetipos_analitico,
            guia_arquetipos=guia_arquetipos,
            dados_microambiente_analitico=dados_microambiente_analitico,
            dados_microambiente_subdimensao=dados_microambiente_subdimensao,
            dados_microambiente_termometro_gaps=dados_microambiente_termometro_gaps,
            dados_microambiente_waterfall_gaps=dados_microambiente_waterfall_gaps,
            guia_microambiente=guia_microambiente
        )

        
        response = jsonify({
            "status": "ok",
            "resposta": resposta_ia
        })

        
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro no chat Leadertrack:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


def buscar_todos_consolidados_leadertrack_rodada(tabela, codrodada, limite=500):
    if not SUPABASE_REST_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise RuntimeError("Supabase service role nao configurado no ambiente.")

    rows = []
    offset = 0
    while True:
        response = requests.get(
            f"{SUPABASE_REST_URL}/{tabela}",
            headers=supabase_headers(prefer_return=False, use_service_role=True),
            params={
                "select": "*",
                "codrodada": f"ilike.{codrodada}",
                "order": "id.asc",
                "limit": str(limite),
                "offset": str(offset),
            },
            timeout=60,
        )
        if response.status_code >= 300:
            raise RuntimeError(
                f"Erro ao consultar {tabela}: HTTP {response.status_code} - {response.text}"
            )
        page = response.json() or []
        rows.extend(page)
        if len(page) < limite:
            break
        offset += limite
    return rows


def calcular_saude_snapshot(registros_arq, registros_micro):
    return calcular_saude_emocional_dashboard(
        registros_arq,
        registros_micro,
        carregar_matriz_arquetipos_rows(),
        carregar_matriz_micro_rows(),
        carregar_saude_emocional_rows(),
    )


def chave_snapshot_executivo(codrodada, scope, source_hash):
    scope_type = str(scope.get("tipo") or "").strip().lower()
    scope_key = str(
        scope.get("empresa")
        or scope.get("holding_id")
        or scope.get("contexto_nome")
        or "contexto"
    ).strip().lower()
    raw = f"{SNAPSHOT_SCHEMA_VERSION}|{codrodada.lower()}|{scope_type}|{scope_key}|{source_hash}"
    return raw[:500]


def resumo_snapshot_executivo(snapshot):
    health = snapshot.get("health") or {}
    scope = snapshot.get("scope") or {}
    return {
        "tipo": scope.get("tipo"),
        "empresa": scope.get("empresa"),
        "contexto_nome": scope.get("contexto_nome"),
        "status": snapshot.get("status"),
        "amostra": snapshot.get("sample"),
        "score_saude_emocional": health.get("score_final"),
        "recortes_elegiveis": len(snapshot.get("cuts") or []),
        "findings_mecanicos": len(snapshot.get("findings") or []),
        "hash_origem": snapshot.get("source_hash"),
    }


def construir_snapshots_executivos_rodada(codrodada, empresas_contexto=None, contexto=None, n_minimo=5):
    rows_arq = buscar_todos_consolidados_leadertrack_rodada("consolidado_arquetipos", codrodada)
    rows_micro = buscar_todos_consolidados_leadertrack_rodada("consolidado_microambiente", codrodada)
    arq_by_company = group_source_rows_by_company(rows_arq)
    micro_by_company = group_source_rows_by_company(rows_micro)
    companies = sorted(set(arq_by_company) | set(micro_by_company))
    snapshots = []

    for company in companies:
        company_arq_rows = arq_by_company.get(company) or []
        company_micro_rows = micro_by_company.get(company) or []
        snapshots.append(build_scope_snapshot(
            scope={"tipo": "empresa", "empresa": company, "codrodada": codrodada},
            archetype_records=registros_arquetipos_consolidados(company_arq_rows),
            microenvironment_records=registros_micro_consolidados(company_micro_rows),
            archetype_rows=company_arq_rows,
            microenvironment_rows=company_micro_rows,
            health_calculator=calcular_saude_snapshot,
            leadertrack_summarizer=resumir_recorte_leadertrack,
            minimum_sample=n_minimo,
            include_cuts=False,
        ))

    selected_companies = {
        str(item or "").strip().lower()
        for item in (empresas_contexto or [])
        if str(item or "").strip()
    }
    selected_companies &= set(companies)
    if selected_companies:
        context_arq_rows = [row for company in sorted(selected_companies) for row in arq_by_company.get(company, [])]
        context_micro_rows = [row for company in sorted(selected_companies) for row in micro_by_company.get(company, [])]
        context_scope = {
            "tipo": "contexto",
            "codrodada": codrodada,
            "empresas": sorted(selected_companies),
            **(contexto or {}),
        }
        snapshots.append(build_scope_snapshot(
            scope=context_scope,
            archetype_records=registros_arquetipos_consolidados(context_arq_rows),
            microenvironment_records=registros_micro_consolidados(context_micro_rows),
            archetype_rows=context_arq_rows,
            microenvironment_rows=context_micro_rows,
            health_calculator=calcular_saude_snapshot,
            leadertrack_summarizer=resumir_recorte_leadertrack,
            minimum_sample=n_minimo,
            include_cuts=True,
            max_cuts=40,
        ))

    return {
        "codrodada": codrodada,
        "empresas_encontradas": companies,
        "fontes": {
            "consolidados_arquetipos": len(rows_arq),
            "consolidados_microambiente": len(rows_micro),
        },
        "snapshots": snapshots,
    }


def persistir_snapshots_executivos(resultado, solicitado_por=None, contexto=None):
    codrodada = resultado["codrodada"]
    now_iso = datetime.utcnow().isoformat(timespec="microseconds") + "Z"
    execution_key = f"{SNAPSHOT_SCHEMA_VERSION}|{codrodada.lower()}|{now_iso}"
    execution = supabase_insert("leadertrack_execucoes_organizacionais", {
        "chave_analise": execution_key,
        "codrodada": codrodada,
        "nivel_contexto": "rodada_todas_empresas",
        "contexto": contexto or {},
        "filtros": {},
        "parametros": {
            "empresas_encontradas": resultado.get("empresas_encontradas") or [],
            "fontes": resultado.get("fontes") or {},
        },
        "status": "processando",
        "versao_regras": SNAPSHOT_SCHEMA_VERSION,
        "solicitado_por": solicitado_por,
        "iniciado_em": now_iso,
    })
    execution_id = execution.get("id")
    saved = []
    try:
        for snapshot in resultado.get("snapshots") or []:
            scope = snapshot.get("scope") or {}
            analysis_key = chave_snapshot_executivo(
                codrodada,
                scope,
                snapshot.get("source_hash") or "sem-hash",
            )
            row = supabase_upsert(
                "leadertrack_pacotes_organizacionais",
                {
                    "execucao_id": execution_id,
                    "chave_analise": analysis_key,
                    "codrodada": codrodada,
                    "nivel_contexto": scope.get("tipo") or "contexto",
                    "empresa_codigo": scope.get("empresa"),
                    "contexto": scope,
                    "filtros": {},
                    "amostra": snapshot.get("sample") or {},
                    "status": snapshot.get("status") or "concluido",
                    "pacote_completo": snapshot,
                    "hash_origem": snapshot.get("source_hash"),
                    "versao_regras": SNAPSHOT_SCHEMA_VERSION,
                    "atualizado_em": now_iso,
                },
                "chave_analise",
            )
            saved.append(row.get("id"))
        supabase_patch("leadertrack_execucoes_organizacionais", execution_id, {
            "status": "concluida",
            "concluido_em": datetime.utcnow().isoformat(timespec="microseconds") + "Z",
            "atualizado_em": datetime.utcnow().isoformat(timespec="microseconds") + "Z",
        })
    except Exception as exc:
        supabase_patch("leadertrack_execucoes_organizacionais", execution_id, {
            "status": "erro",
            "erro_resumido": str(exc)[:1000],
            "concluido_em": datetime.utcnow().isoformat(timespec="microseconds") + "Z",
            "atualizado_em": datetime.utcnow().isoformat(timespec="microseconds") + "Z",
        })
        raise
    return {"execucao_id": execution_id, "pacotes_salvos": len(saved)}


def autorizado_snapshot_executivo():
    configured = str(LEADERTRACK_SNAPSHOT_ADMIN_KEY or "")
    provided = str(request.headers.get("X-HRKey-Snapshot-Key") or "")
    return bool(configured and provided and hmac.compare_digest(configured, provided))


def buscar_snapshots_contexto_rodada(codrodada, limite=100):
    if not SUPABASE_REST_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise RuntimeError("Supabase service role nao configurado no ambiente.")
    response = requests.get(
        f"{SUPABASE_REST_URL}/leadertrack_pacotes_organizacionais",
        headers=supabase_headers(prefer_return=False, use_service_role=True),
        params={
            "select": "id,codrodada,nivel_contexto,contexto,amostra,status,pacote_completo,versao_regras,gerado_em,atualizado_em",
            "codrodada": f"ilike.{codrodada}",
            "nivel_contexto": "eq.contexto",
            "order": "atualizado_em.desc",
            "limit": str(limite),
        },
        timeout=60,
    )
    if response.status_code >= 300:
        raise RuntimeError(
            "Erro ao consultar snapshots executivos: "
            f"HTTP {response.status_code} - {response.text}"
        )
    return response.json() or []


@app.route("/buscar-snapshot-executivo-leadertrack", methods=["POST", "OPTIONS"])
def buscar_snapshot_executivo_leadertrack():
    if request.method == "OPTIONS":
        return jsonify({"status": "CORS preflight OK"}), 200
    if not LEADERTRACK_SNAPSHOT_ADMIN_KEY:
        return jsonify({"erro": "Chave de snapshots nao configurada no backend."}), 503
    if not autorizado_snapshot_executivo():
        return jsonify({"erro": "Leitura executiva nao autorizada."}), 403

    dados = request.get_json() or {}
    codrodada = str(dados.get("codrodada") or "").strip()
    contexto_solicitado = {
        key: dados.get(key)
        for key in (
            "cliente_id", "holding_id", "contexto_nome", "holding_nome", "contexto"
        )
        if dados.get(key) not in (None, "")
    }
    if not codrodada:
        return jsonify({"erro": "Informe a rodada do snapshot."}), 400
    if not contexto_solicitado:
        return jsonify({"erro": "Contexto obrigatorio para consultar o snapshot."}), 400

    try:
        rows = buscar_snapshots_contexto_rodada(codrodada)
        for row in rows:
            package = row.get("pacote_completo") or {}
            if not isinstance(package, dict):
                continue
            if not snapshot_matches_context(package, contexto_solicitado):
                continue
            return jsonify({
                "status": "ok",
                "snapshot": snapshot_for_frontend(package),
                "metadados": {
                    "pacote_id": row.get("id"),
                    "codrodada": row.get("codrodada"),
                    "versao_regras": row.get("versao_regras"),
                    "gerado_em": row.get("gerado_em"),
                    "atualizado_em": row.get("atualizado_em"),
                },
            }), 200
        return jsonify({
            "erro": "Nenhum snapshot executivo encontrado para esta rodada e contexto.",
            "codrodada": codrodada,
        }), 404
    except Exception as exc:
        print("Erro ao buscar snapshot executivo LeaderTrack:", exc)
        return jsonify({"erro": str(exc), "status": "erro_leitura_snapshot"}), 500


@app.route("/gerar-snapshots-executivos-leadertrack", methods=["POST", "OPTIONS"])
def gerar_snapshots_executivos_leadertrack():
    if request.method == "OPTIONS":
        return jsonify({"status": "CORS preflight OK"}), 200
    if not LEADERTRACK_SNAPSHOT_ADMIN_KEY:
        return jsonify({"erro": "Chave administrativa de snapshots nao configurada no backend."}), 503
    if not autorizado_snapshot_executivo():
        return jsonify({"erro": "Acao administrativa nao autorizada."}), 403

    dados = request.get_json() or {}
    codrodada = str(dados.get("codrodada") or "").strip()
    if not codrodada:
        return jsonify({"erro": "Informe a rodada a processar."}), 400
    try:
        n_minimo = max(3, min(50, int(dados.get("nMinimo") or dados.get("n_minimo") or 5)))
    except Exception:
        n_minimo = 5
    contexto = {
        key: dados.get(key)
        for key in (
            "cliente_id", "holding_id", "empresa_id", "filial_id",
            "nivel_contexto", "contexto_nome", "holding_nome",
        )
        if dados.get(key) not in (None, "")
    }
    try:
        resultado = construir_snapshots_executivos_rodada(
            codrodada,
            empresas_contexto=dados.get("empresasContexto") or dados.get("empresas_contexto") or [],
            contexto=contexto,
            n_minimo=n_minimo,
        )
        if not resultado.get("empresas_encontradas"):
            return jsonify({
                "erro": "Nenhuma empresa com consolidados foi encontrada nesta rodada.",
                "codrodada": codrodada,
            }), 404

        persistir = bool_param(dados.get("persistir"), True)
        persistence = None
        if persistir:
            persistence = persistir_snapshots_executivos(
                resultado,
                solicitado_por=request.headers.get("X-HRKey-Admin-User"),
                contexto=contexto,
            )
        summaries = [resumo_snapshot_executivo(item) for item in resultado.get("snapshots") or []]
        return jsonify({
            "status": "concluido" if persistir else "previsualizacao_sem_gravacao",
            "codrodada": codrodada,
            "empresas_encontradas": resultado.get("empresas_encontradas"),
            "quantidade_empresas": len(resultado.get("empresas_encontradas") or []),
            "fontes": resultado.get("fontes"),
            "snapshots": summaries,
            "persistencia": persistence,
            "versao_regras": SNAPSHOT_SCHEMA_VERSION,
        }), 200
    except Exception as exc:
        print("Erro ao gerar snapshots executivos LeaderTrack:", exc)
        return jsonify({"erro": str(exc), "status": "erro_snapshots"}), 500


@app.route("/previsualizar-saude-emocional-leadertrack", methods=["POST", "OPTIONS"])
def previsualizar_saude_emocional_leadertrack():
    """Executa somente a previsualizacao rastreavel do snapshot executivo."""
    if request.method == "OPTIONS":
        response = jsonify({"status": "CORS preflight OK"})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        empresa = str(dados.get("empresa") or dados.get("company") or "").strip().lower()
        codrodada = str(dados.get("codrodada") or "").strip().lower()
        contexto_ids = {
            "cliente_id": dados.get("cliente_id") or dados.get("clienteId"),
            "holding_id": dados.get("holding_id") or dados.get("holdingId"),
            "empresa_id": dados.get("empresa_id") or dados.get("empresaId"),
            "filial_id": dados.get("filial_id") or dados.get("filialId"),
        }
        empresas_contexto = dados.get("empresasContexto") or dados.get("empresas_contexto") or []

        if not codrodada or (not empresa and not any(contexto_ids.values())):
            return jsonify({
                "erro": "Informe rodada e empresa ou identificadores de contexto.",
                "campos_necessarios": ["codrodada", "empresa ou contexto_ids"],
            }), 400

        # Quando a tela envia a lista de empresas do contexto, ela e a fonte
        # explicita do escopo. O dashboard original da PROSPERA tambem consulta
        # os consolidados pelas empresas, sem restringir antes pelos IDs.
        contexto_consulta = {} if empresas_contexto else contexto_ids
        arq_consolidados = buscar_consolidados_leadertrack_contexto(
            "consolidado_arquetipos", empresa, codrodada, contexto_consulta,
            empresas_contexto=empresas_contexto,
        )
        micro_consolidados = buscar_consolidados_leadertrack_contexto(
            "consolidado_microambiente", empresa, codrodada, contexto_consulta,
            empresas_contexto=empresas_contexto,
        )
        registros_arq = registros_arquetipos_consolidados(arq_consolidados)
        registros_micro = registros_micro_consolidados(micro_consolidados)
        resultado = calcular_saude_emocional_dashboard(
            registros_arq,
            registros_micro,
            carregar_matriz_arquetipos_rows(),
            carregar_matriz_micro_rows(),
            carregar_saude_emocional_rows(),
        )
        if not resultado.get("quantidade_afirmacoes_calculadas"):
            return jsonify({
                "erro": "Nao houve dados suficientes para calcular a previsualizacao de saude emocional.",
                "amostra": {
                    "respondentes_arquetipos": resultado.get("respondentes_arquetipos"),
                    "respondentes_microambiente": resultado.get("respondentes_microambiente"),
                    "afirmacoes_calculadas": 0,
                },
            }), 404

        lideres = {
            registro.get("email_lider")
            for registro in registros_arq + registros_micro
            if registro.get("email_lider")
        }
        response = jsonify({
            "status": "previsualizacao_sem_gravacao",
            "escopo": "executivo_agregado",
            "saude_emocional": resultado,
            "amostra": {
                "lideres": len(lideres),
                "respondentes_arquetipos": resultado.get("respondentes_arquetipos"),
                "respondentes_microambiente": resultado.get("respondentes_microambiente"),
                "afirmacoes_calculadas": resultado.get("quantidade_afirmacoes_calculadas"),
            },
            "rastreabilidade": {
                "fonte_arquetipos": "consolidado_arquetipos por lider",
                "fonte_microambiente": "consolidado_microambiente por lider",
                "projeto_supabase": (SUPABASE_REST_URL or "").split("/")[2] if "/" in (SUPABASE_REST_URL or "") else None,
                "regra": "mesma sequencia do dashboard: matriz por respondente, media por afirmacao e media das cinco dimensoes",
                "gravacao_realizada": False,
                "ia_chamada": False,
            },
        })
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200
    except Exception as e:
        print("Erro na previsualizacao de saude emocional LeaderTrack:", e)
        response = jsonify({"erro": str(e), "status": "erro_previsualizacao"})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/gerar-devolutiva-leadertrack", methods=["POST", "OPTIONS"])
def gerar_devolutiva_leadertrack():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        empresa = str(dados.get("empresa") or dados.get("company") or "").strip().lower()
        codrodada = str(dados.get("codrodada") or "").strip().lower()
        email_lider = str(dados.get("emailLider") or "").strip().lower()
        contexto = dados.get("contexto", "")
        equipe_tipo = str(dados.get("equipeTipo") or dados.get("tipoEquipe") or "direta").strip().lower()
        contexto_ids = {
            "cliente_id": dados.get("cliente_id") or dados.get("clienteId"),
            "holding_id": dados.get("holding_id") or dados.get("holdingId"),
            "empresa_id": dados.get("empresa_id") or dados.get("empresaId"),
            "filial_id": dados.get("filial_id") or dados.get("filialId"),
        }
        empresas_contexto = dados.get("empresasContexto") or dados.get("empresas_contexto") or []
        nome_lider = dados.get("nomeLider", "")
        gap_minimo = float(dados.get("gapMinimo", 20) or 20)
        baixa_referencia_threshold = float(dados.get("baixaReferenciaThreshold", 70) or 70)
        limite_gaps = dados.get("limiteGaps")
        limite_gaps = int(limite_gaps) if limite_gaps not in (None, "", 0, "0") else None
        maximo_gaps_por_ciclo = int(dados.get("maximoGapsPorCiclo", 4) or 4)
        gerar_apenas_primeiro_ciclo = bool_param(dados.get("gerarApenasPrimeiroCiclo"), True)
        gerar_planos_com_ia = bool_param(dados.get("gerarPlanosComIA"), False)
        persistir = bool_param(dados.get("persistir"), False)
        incluir_guias_caderno = bool_param(
            dados.get("incluirGuiasCaderno") or dados.get("incluir_guias_caderno"),
            False,
        )
        gerado_por = dados.get("geradoPor")
        indicadores_disponiveis = dados.get("indicadoresOperacionaisDisponiveis", [])
        considerar_todos_lideres = (
            bool_param(dados.get("considerarTodosLideres"), False)
            or bool_param(dados.get("todosLideresContexto"), False)
            or leadertrack_todos_lideres(email_lider)
        )

        if considerar_todos_lideres and not empresa:
            empresa = "todos"
        if considerar_todos_lideres:
            email_lider = "__todos_lideres_contexto__"
            nome_lider = nome_lider or "Todos os lideres do contexto"

        if not codrodada or not email_lider or (
            not empresa and not any(contexto_ids.values())
        ):
            response = jsonify({
                "erro": "Campos obrigatorios ausentes.",
                "campos_necessarios": ["empresa ou contexto_ids", "codrodada", "emailLider"]
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        if considerar_todos_lideres and persistir:
            response = jsonify({
                "erro": "Persistencia bloqueada para devolutiva consolidada de todos os lideres.",
                "orientacao": (
                    "A acao consolidada gera a mesma devolutiva em modo contexto, mas nao grava "
                    "PDI/meta individual automaticamente. Gere com persistir=false, valide a leitura "
                    "e depois desdobre em acoes individuais quando necessario."
                ),
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        if considerar_todos_lideres and gerar_planos_com_ia:
            response = jsonify({
                "erro": "Geracao profunda de PDI com IA bloqueada para todos os lideres.",
                "orientacao": (
                    "Use a devolutiva consolidada para diagnostico do contexto. PDIs profundos "
                    "continuam sendo gerados por lider/afirmacao para evitar plano individual indevido."
                ),
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        if persistir and not gerar_planos_com_ia:
            response = jsonify({
                "erro": "Persistencia bloqueada para devolutiva sem planos gerados pela IA.",
                "orientacao": "Envie gerarPlanosComIA=true apenas quando for gerar e salvar um PDI validado."
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        prompt_base = carregar_prompt_leadertrack() if gerar_planos_com_ia else ""

        if considerar_todos_lideres:
            inputs_consolidados = buscar_inputs_devolutiva_todos_lideres(
                empresa,
                codrodada,
                contexto_ids,
                empresas_contexto=empresas_contexto,
            )
            dados_arquetipos_comparativo = inputs_consolidados["dados_arquetipos_comparativo"]
            dados_arquetipos_analitico = inputs_consolidados["dados_arquetipos_analitico"]
            guia_arquetipos = inputs_consolidados["guia_arquetipos"]
            dados_microambiente_analitico = inputs_consolidados["dados_microambiente_analitico"]
            dados_microambiente_auto_dimensao = inputs_consolidados["dados_microambiente_auto_dimensao"]
            dados_microambiente_auto_subdimensao = inputs_consolidados["dados_microambiente_auto_subdimensao"]
            dados_microambiente_media_dimensao = inputs_consolidados["dados_microambiente_media_dimensao"]
            dados_microambiente_subdimensao = inputs_consolidados["dados_microambiente_subdimensao"]
            dados_microambiente_termometro_gaps = inputs_consolidados["dados_microambiente_termometro_gaps"]
            dados_microambiente_waterfall_gaps = inputs_consolidados["dados_microambiente_waterfall_gaps"]
            guia_microambiente = inputs_consolidados["guia_microambiente"]
            metadados_consolidado = inputs_consolidados["metadados"]
            camada_executiva = inputs_consolidados.get("devolutiva_executiva") or {}
        else:
            metadados_consolidado = {}
            camada_executiva = {}
            dados_arquetipos_comparativo = buscar_json_supabase(
                "arquetipos_grafico_comparativo", empresa, codrodada, email_lider
            )
            dados_arquetipos_analitico = buscar_json_supabase(
                "arquetipos_analitico", empresa, codrodada, email_lider
            )
            guia_arquetipos = buscar_json_supabase(
                "arquetipos_parecer_ia", empresa, codrodada, email_lider
            )
            dados_microambiente_analitico = buscar_json_microambiente(
                "microambiente_analitico", empresa, codrodada, email_lider
            )
            dados_microambiente_auto_dimensao = buscar_json_microambiente(
                "microambiente_grafico_autoavaliacao_dimensao", empresa, codrodada, email_lider
            )
            dados_microambiente_auto_subdimensao = buscar_json_microambiente(
                "microambiente_grafico_autoavaliacao_subdimensao", empresa, codrodada, email_lider
            )
            dados_microambiente_media_dimensao = buscar_json_microambiente(
                "microambiente_grafico_mediaequipe_dimensao", empresa, codrodada, email_lider
            )
            dados_microambiente_subdimensao = buscar_json_microambiente(
                "microambiente_grafico_mediaequipe_subdimensao", empresa, codrodada, email_lider
            )
            dados_microambiente_termometro_gaps = buscar_json_microambiente(
                "microambiente_termometro_gaps", empresa, codrodada, email_lider
            )
            dados_microambiente_waterfall_gaps = buscar_json_microambiente(
                "microambiente_waterfall_gaps", empresa, codrodada, email_lider
            )
            guia_microambiente = buscar_json_microambiente(
                "microambiente_parecer_ia", empresa, codrodada, email_lider
            )

        if considerar_todos_lideres and (
            int(metadados_consolidado.get("lideres_com_arquetipos") or 0) == 0
            or int(metadados_consolidado.get("lideres_com_microambiente") or 0) == 0
        ):
            response = jsonify({
                "erro": "Dados LeaderTrack insuficientes para gerar devolutiva consolidada.",
                "escopo": "todos_lideres_contexto",
                "dados_encontrados": metadados_consolidado,
                "orientacao": (
                    "Verifique se a rodada e o contexto selecionado possuem relatorios individuais "
                    "ja calculados para arquetipos e microambiente."
                ),
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 404

        if not dados_arquetipos_comparativo or not dados_microambiente_analitico:
            response = jsonify({
                "erro": "Dados LeaderTrack insuficientes para gerar devolutiva.",
                "dados_encontrados": {
                    "arquetipos_grafico_comparativo": bool(dados_arquetipos_comparativo),
                    "microambiente_analitico": bool(dados_microambiente_analitico),
                }
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 404

        arquetipos = archetype_summary(dados_arquetipos_comparativo)
        todas_afirmacoes = microenvironment_affirmations(dados_microambiente_analitico)
        gaps = filter_gaps(todas_afirmacoes, gap_minimo)
        if limite_gaps:
            gaps = gaps[:limite_gaps]
        baixa_referencia = low_reference_affirmations(todas_afirmacoes, baixa_referencia_threshold)
        amostra = avaliar_amostra_leadertrack(
            dados_arquetipos_comparativo,
            dados_arquetipos_analitico,
            dados_microambiente_analitico,
            dados_microambiente_subdimensao,
            dados_microambiente_termometro_gaps,
            dados_microambiente_waterfall_gaps,
        )
        if considerar_todos_lideres:
            amostra["escopo"] = "todos_lideres_contexto"
            amostra["fonte_arquetipos"] = metadados_consolidado.get("fonte_arquetipos")
            amostra["fonte_microambiente"] = metadados_consolidado.get("fonte_microambiente")
            amostra["lideres_com_arquetipos"] = metadados_consolidado.get("lideres_com_arquetipos")
            amostra["lideres_com_microambiente"] = metadados_consolidado.get("lideres_com_microambiente")
            amostra["respostas_arquetipos_equipe"] = metadados_consolidado.get("respostas_arquetipos_equipe")
            amostra["respostas_microambiente_equipe"] = metadados_consolidado.get("respostas_microambiente_equipe")
            amostra["saude_emocional_score"] = (camada_executiva.get("saude_emocional_geral") or {}).get("score_final")
            amostra["saude_emocional_label"] = (camada_executiva.get("saude_emocional_geral") or {}).get("label")
            amostra["recortes_executivos"] = len(camada_executiva.get("recortes") or [])
            amostra["lideres"] = max(
                int(metadados_consolidado.get("lideres_com_arquetipos") or 0),
                int(metadados_consolidado.get("lideres_com_microambiente") or 0),
            )
        gaps_para_gerar = gaps[:maximo_gaps_por_ciclo] if gerar_apenas_primeiro_ciclo else gaps
        leader = {
            "nome": nome_lider,
            "email": email_lider,
            "rodada": codrodada,
            "contexto": contexto,
            "contexto_ids": contexto_ids,
        }
        devolutiva = build_empty_devolutiva(
            empresa=empresa,
            contexto=contexto,
            codrodada=codrodada,
            email_lider=email_lider,
            nome_lider=nome_lider,
            gaps=gaps,
            arquetipos=arquetipos,
            maximo_gaps_por_ciclo=maximo_gaps_por_ciclo,
            todas_afirmacoes=todas_afirmacoes,
            baixa_referencia=baixa_referencia,
            gap_minimo=gap_minimo,
            baixa_referencia_threshold=baixa_referencia_threshold,
            contexto_ids=contexto_ids,
        )
        devolutiva["equipe_tipo"] = equipe_tipo
        devolutiva["amostra"] = amostra
        if considerar_todos_lideres:
            devolutiva["escopo_devolutiva"] = {
                "tipo": "todos_lideres_contexto",
                "descricao": "Mesma devolutiva LeaderTrack, consolidada para todos os lideres e respondentes do contexto selecionado.",
                "persistencia_automatica": False,
                "pdi_individual_automatico": False,
            }
            devolutiva["devolutiva_executiva"] = camada_executiva
            devolutiva["avisos"] = (devolutiva.get("avisos") or []) + [
                "Devolutiva consolidada: representa todos os lideres do contexto selecionado, nao um lider individual.",
                "Persistencia, PDI individual e meta de desempenho automatica ficam bloqueados neste modo.",
            ]
        if amostra.get("insuficiente"):
            devolutiva["modo_devolutiva"] = {
                "modo": "amostra_insuficiente",
                "titulo": "Amostra insuficiente para leitura coletiva robusta",
                "leitura": amostra.get("orientacao"),
            }
            devolutiva["avisos"] = (devolutiva.get("avisos") or []) + [
                "Amostra insuficiente: tratar graficos e planos como leitura limitada, sem conclusoes fortes sobre a equipe.",
            ]
        devolutiva["status"] = "gerada_sem_persistencia"
        devolutiva["modo_geracao_planos"] = "com_ia" if gerar_planos_com_ia else "estrutura_sem_ia"
        if incluir_guias_caderno:
            devolutiva["guias_caderno"] = {
                "arquetipos": guia_caderno_payload(
                    "arquetipos",
                    guia_arquetipos,
                    {
                        "comparativo": dados_arquetipos_comparativo,
                        "analitico": dados_arquetipos_analitico,
                    },
                    {"amostra": amostra},
                ),
                "microambiente": guia_caderno_payload(
                    "microambiente",
                    guia_microambiente,
                    {
                        "analitico": dados_microambiente_analitico,
                        "autoavaliacao_dimensao": dados_microambiente_auto_dimensao,
                        "autoavaliacao_subdimensao": dados_microambiente_auto_subdimensao,
                        "mediaequipe_dimensao": dados_microambiente_media_dimensao,
                        "mediaequipe_subdimensao": dados_microambiente_subdimensao,
                        "termometro_gaps": dados_microambiente_termometro_gaps,
                        "waterfall_gaps": dados_microambiente_waterfall_gaps,
                    },
                    {"amostra": amostra},
                ),
            }
        devolutiva["proximo_passo_sugerido"] = (
            "Gerar PDI detalhado por afirmacao, em chamada especifica com IA, para evitar timeout."
            if not gerar_planos_com_ia else
            "Validar plano gerado antes de enviar para PDI/Treinamentos ou Desempenho."
        )

        for gap in gaps_para_gerar:
            gap_id = f"{gap['questao']}_{slug(gap['dimensao'])}_{slug(gap['subdimensao'])}"
            if gerar_planos_com_ia:
                prompt_diagnostico = build_diagnostic_prompt(
                    leader=leader,
                    arquetipos=arquetipos,
                    gap=gap,
                    indicadores_disponiveis=indicadores_disponiveis,
                )
                resposta_diagnostico = gerar_resposta_ia_leadertrack(
                    pergunta=prompt_diagnostico,
                    prompt_base=prompt_base,
                    empresa=empresa,
                    codrodada=codrodada,
                    email_lider=email_lider,
                    pagina_atual="/gerar-devolutiva-leadertrack",
                    url_atual="https://gestor.thehrkey.tech",
                    dados_arquetipos_comparativo=dados_arquetipos_comparativo,
                    dados_arquetipos_analitico=dados_arquetipos_analitico,
                    guia_arquetipos=guia_arquetipos,
                    dados_microambiente_analitico=dados_microambiente_analitico,
                    dados_microambiente_subdimensao=dados_microambiente_subdimensao,
                    dados_microambiente_termometro_gaps=dados_microambiente_termometro_gaps,
                    dados_microambiente_waterfall_gaps=dados_microambiente_waterfall_gaps,
                    guia_microambiente=guia_microambiente,
                )
                diagnostico = parse_json_response(resposta_diagnostico)

                semanas = []
                revisoes = []
                for inicio, fim in [(1, 4), (5, 8), (9, 12)]:
                    prompt_semanal = build_weekly_prompt(
                        leader=leader,
                        arquetipos=arquetipos,
                        gap=gap,
                        diagnostic=diagnostico,
                        start_week=inicio,
                        end_week=fim,
                        indicadores_disponiveis=indicadores_disponiveis,
                    )
                    resposta_semanal = gerar_resposta_ia_leadertrack(
                        pergunta=prompt_semanal,
                        prompt_base=prompt_base,
                        empresa=empresa,
                        codrodada=codrodada,
                        email_lider=email_lider,
                        pagina_atual="/gerar-devolutiva-leadertrack",
                        url_atual="https://gestor.thehrkey.tech",
                        dados_arquetipos_comparativo=dados_arquetipos_comparativo,
                        dados_arquetipos_analitico=dados_arquetipos_analitico,
                        guia_arquetipos=guia_arquetipos,
                        dados_microambiente_analitico=dados_microambiente_analitico,
                        dados_microambiente_subdimensao=dados_microambiente_subdimensao,
                        dados_microambiente_termometro_gaps=dados_microambiente_termometro_gaps,
                        dados_microambiente_waterfall_gaps=dados_microambiente_waterfall_gaps,
                        guia_microambiente=guia_microambiente,
                    )
                    plano = parse_json_response(resposta_semanal)
                    semanas.extend(plano.get("plano_12_semanas", []))
                    if plano.get("revisao_parcial_informal"):
                        revisoes.append(plano["revisao_parcial_informal"])

                plano_12_semanas = sorted(semanas, key=lambda item: int(item.get("semana", 0) or 0))
            else:
                diagnostico = {
                    "status": "pendente_geracao_ia",
                    "mensagem": "Diagnostico tecnico e plano semanal devem ser gerados por etapa para evitar timeout.",
                    "gap": gap,
                    "arquetipos_disponiveis_para_cruzamento": arquetipos,
                    "indicadores_operacionais_disponiveis": indicadores_disponiveis,
                }
                plano_12_semanas = []
                revisoes = []

            pdi_payload = {
                "gap_id": gap_id,
                "gap": gap,
                "diagnostico": diagnostico,
                "plano_12_semanas": plano_12_semanas,
                "revisoes_parciais_informais": revisoes,
                "geracao_ia": {
                    "status": "concluida" if gerar_planos_com_ia else "pendente",
                    "motivo": None if gerar_planos_com_ia else "A tela deve solicitar a geracao profunda deste gap em chamada propria.",
                },
                "origem_para_pdi_treinamentos": {
                    "pode_enviar_para_modulo_pdi": bool(gerar_planos_com_ia),
                    "status_inicial": "sugerido_pela_devolutiva",
                    "requer_validacao_consultiva": True,
                },
            }
            historico_snapshot = dict(pdi_payload)
            pdi_payload["historico_evento_inicial"] = build_history_event(
                empresa=empresa,
                contexto=contexto,
                email_lider=email_lider,
                nome_lider=nome_lider,
                codrodada=codrodada,
                gap_id=gap_id,
                event_type="pdi_sugerido_pela_devolutiva",
                description="PDI sugerido a partir de devolutiva LeaderTrack, pendente de validacao consultiva.",
                payload=historico_snapshot,
            )
            pdi_payload["meta_desempenho_sugerida"] = build_performance_goal_suggestion(
                email_lider=email_lider,
                empresa=empresa,
                contexto=contexto,
                codrodada=codrodada,
                gap_id=gap_id,
                gap=gap,
            )
            devolutiva["pdis"].append(pdi_payload)

        if persistir:
            persistencia = persistir_devolutiva_leadertrack(devolutiva, gerado_por=gerado_por)
            devolutiva["persistencia"] = persistencia
            devolutiva["status"] = "gerada_e_salva_como_rascunho"

        response = jsonify(devolutiva)
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro ao gerar devolutiva LeaderTrack:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/gerar-devolutiva-organizacional-leadertrack", methods=["POST", "OPTIONS"])
def gerar_devolutiva_organizacional_leadertrack():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        pacote = dados.get("pacote_analitico") or dados.get("pacoteAnalitico") or {}
        gerar_com_ia = bool_param(dados.get("gerarComIA"), False)
        persistir = bool_param(dados.get("persistir"), False)

        valido, erro = validate_organizational_package(pacote)
        if not valido:
            response = jsonify({
                "erro": erro,
                "status": "pacote_organizacional_invalido",
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        if persistir:
            response = jsonify({
                "erro": "Persistencia da devolutiva organizacional ainda nao esta habilitada.",
                "status": "persistencia_bloqueada",
                "orientacao": (
                    "Primeiro valide a geracao sem gravacao. Depois crie a tabela/cache "
                    "organizacional no Supabase com RLS e regra de acesso propria."
                ),
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        prompt_organizacional = build_organizational_feedback_prompt(pacote)
        resposta_base = {
            "status": "preparada_sem_ia" if not gerar_com_ia else "gerada_com_ia",
            "tipo": "devolutiva_organizacional_leadertrack",
            "contexto": pacote.get("contexto") or {},
            "filtros": pacote.get("filtros") or {},
            "amostra": pacote.get("amostra") or {},
            "governanca": pacote.get("governanca") or {},
            "achados_relevantes_recebidos": len(pacote.get("achados_relevantes") or []),
            "geracao_ia": {
                "solicitada": gerar_com_ia,
                "executada": False,
            },
        }

        if not gerar_com_ia:
            resposta_base["proximo_passo"] = (
                "Enviar gerarComIA=true quando o pacote analitico estiver validado na tela."
            )
            response = jsonify(resposta_base)
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 200

        try:
            prompt_base = (
                "Voce e o LeaderTrackbot em modo de parecer corporativo para RH, CEO e diretoria. "
                "Use somente o pacote analitico enviado pelo usuario. Nao invente dados, percentuais, "
                "nomes, causas, historico, modalidade de trabalho, politicas internas, turnover, retencao ou cruzamentos. "
                "Copie n, score, real, ideal, gap e delta exatamente do pacote, sem recalcular. "
                "Nao use recomendacoes genericas se elas nao nascerem de achado medido. "
                "Evite palavras como intervencao, workshop, treinamento generico, team building e taxa de resposta "
                "quando o pacote nao trouxer essas evidencias. "
                "Saude emocional deve ser tratada somente "
                "em nivel organizacional/agregado e nunca como devolutiva individual para lider. "
                "Nao faca diagnostico clinico nem atribua culpa a lideres ou grupos. Use Daniel Goleman/HBR apenas como "
                "apoio conceitual prudente sobre lideranca, clima emocional e inteligencia emocional, "
                "sem citacao textual e sem substituir o modelo LeaderTrack. Responda somente JSON valido."
            )
            modelo_ia = str(dados.get("modelo") or "gpt-4o-mini")
            resposta_ia = gerar_resposta_ia_leadertrack_enxuta(
                pergunta=prompt_organizacional,
                prompt_base=prompt_base,
                model=modelo_ia,
                max_tokens=int(dados.get("maxTokens") or 1400),
                timeout=int(dados.get("timeout") or 15),
                temperature=float(dados.get("temperature") or 0.2),
            )
            resposta_json = parse_json_response(resposta_ia)
            resposta_json = revisar_devolutiva_organizacional_ia(resposta_json, pacote)
        except Exception as erro_ia:
            print("Erro na IA da devolutiva organizacional LeaderTrack:", erro_ia)
            resposta_base["status"] = "preparada_com_erro_ia"
            resposta_base["geracao_ia"] = {
                "solicitada": True,
                "executada": False,
                "erro": str(erro_ia),
            }
            if "resposta_ia" in locals() and resposta_ia:
                resposta_base["devolutiva"] = {
                    "texto_gerado_nao_estruturado": str(resposta_ia),
                    "observacao": (
                        "A IA retornou texto, mas ele nao veio em JSON valido. "
                        "O conteudo bruto foi preservado para evitar perda da analise."
                    ),
                }
            resposta_base["orientacao"] = (
                "O pacote analitico foi recebido e validado, mas a IA nao concluiu a geracao "
                "do texto executivo nesta tentativa. Reenvie a geracao ou reduza filtros/achados."
            )
            response = jsonify(resposta_base)
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 200

        resposta_base["geracao_ia"] = {
            "solicitada": True,
            "executada": True,
            "modelo": modelo_ia,
        }
        resposta_base["devolutiva"] = resposta_json

        response = jsonify(resposta_base)
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro ao gerar devolutiva organizacional LeaderTrack:", e)
        response = jsonify({
            "erro": str(e),
            "status": "erro_geracao_devolutiva_organizacional",
        })
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/listar-empresas-leadertrack", methods=["POST", "OPTIONS"])
def listar_empresas_leadertrack():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        codrodada = str(dados.get("codrodada") or "").strip().lower()
        contexto = dados.get("contexto", "")
        contexto_ids = {
            "cliente_id": dados.get("cliente_id") or dados.get("clienteId"),
            "holding_id": dados.get("holding_id") or dados.get("holdingId"),
            "empresa_id": dados.get("empresa_id") or dados.get("empresaId"),
            "filial_id": dados.get("filial_id") or dados.get("filialId"),
        }

        empresas = listar_empresas_relatorios(codrodada=codrodada or None)
        response = jsonify({
            "status": "ok",
            "codrodada": codrodada,
            "contexto": contexto,
            "contexto_ids": contexto_ids,
            "total": len(empresas),
            "empresas": empresas,
            "observacao": "Lista baseada em empresas tecnicas encontradas nos relatorios LeaderTrack ja apurados. Nome oficial deve ser vinculado ao cadastro central em uma proxima etapa.",
        })
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro ao listar empresas LeaderTrack:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/listar-rodadas-leadertrack", methods=["POST", "OPTIONS"])
def listar_rodadas_leadertrack():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        empresa = str(dados.get("empresa") or "").strip().lower()
        email_lider = str(dados.get("emailLider") or dados.get("email_lider") or "").strip().lower()
        contexto = dados.get("contexto", "")
        contexto_ids = {
            "cliente_id": dados.get("cliente_id") or dados.get("clienteId"),
            "holding_id": dados.get("holding_id") or dados.get("holdingId"),
            "empresa_id": dados.get("empresa_id") or dados.get("empresaId"),
            "filial_id": dados.get("filial_id") or dados.get("filialId"),
        }

        rodadas = listar_rodadas_relatorios(
            empresa=empresa or None,
            email_lider=email_lider or None,
            contexto_ids=contexto_ids,
        )
        response = jsonify({
            "status": "ok",
            "empresa": empresa,
            "email_lider": email_lider,
            "contexto": contexto,
            "contexto_ids": contexto_ids,
            "total": len(rodadas),
            "rodadas": rodadas,
            "observacao": "Lista baseada nas rodadas encontradas nos relatorios LeaderTrack ja apurados.",
        })
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro ao listar rodadas LeaderTrack:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/listar-lideres-leadertrack", methods=["POST", "OPTIONS"])
def listar_lideres_leadertrack():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        empresa = dados.get("empresa", "").lower()
        codrodada = dados.get("codrodada", "").lower()
        contexto = dados.get("contexto", "")
        contexto_ids = {
            "cliente_id": dados.get("cliente_id") or dados.get("clienteId"),
            "holding_id": dados.get("holding_id") or dados.get("holdingId"),
            "empresa_id": dados.get("empresa_id") or dados.get("empresaId"),
            "filial_id": dados.get("filial_id") or dados.get("filialId"),
        }

        if not empresa or not codrodada:
            response = jsonify({
                "erro": "Campos obrigatorios ausentes.",
                "campos_necessarios": ["empresa", "codrodada"]
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        lideres = listar_lideres_relatorios(empresa, codrodada)
        response = jsonify({
            "status": "ok",
            "empresa": empresa,
            "codrodada": codrodada,
            "contexto": contexto,
            "contexto_ids": contexto_ids,
            "total": len(lideres),
            "lideres": lideres,
            "observacao": "Lista baseada em relatorios LeaderTrack ja apurados. Nome completo depende de vinculo futuro com cadastro central.",
        })
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro ao listar lideres LeaderTrack:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/teste-cache-leadertrack", methods=["POST", "OPTIONS"])
def teste_cache_leadertrack():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        empresa = str(dados.get("empresa") or "fastco").strip().lower()
        codrodada = str(dados.get("codrodada") or "av1225").strip().lower()
        email_lider = str(dados.get("emailLider") or "teste-cache@leadertrack.local").strip().lower()
        nome_lider = str(dados.get("nomeLider") or "Teste Cache LeaderTrack").strip()
        contexto = str(dados.get("contexto") or "teste_cache").strip()
        contexto_ids = {
            "cliente_id": dados.get("cliente_id") or dados.get("clienteId"),
            "holding_id": dados.get("holding_id") or dados.get("holdingId"),
            "empresa_id": dados.get("empresa_id") or dados.get("empresaId"),
            "filial_id": dados.get("filial_id") or dados.get("filialId"),
        }
        cache_key = leadertrack_cache_key(
            empresa=empresa,
            codrodada=codrodada,
            email_lider=email_lider,
            equipe_tipo="diagnostico",
            gap_id="teste_cache",
            etapa="teste_cache",
        )
        payload_teste = {
            "status": "ok",
            "etapa": "teste_cache",
            "gap_id": "teste_cache",
            "fonte": "teste_sem_ia",
            "cache": {
                "status": "miss",
                "cache_key": cache_key,
            },
        }
        persistencia = salvar_cache_leadertrack(
            empresa=empresa,
            contexto=contexto,
            contexto_ids=contexto_ids,
            email_lider=email_lider,
            nome_lider=nome_lider,
            cache_key=cache_key,
            payload=payload_teste,
            gerado_por=dados.get("geradoPor"),
        )
        payload_teste["persistencia"] = persistencia
        response = jsonify(payload_teste)
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        status_code = 200 if persistencia.get("status") == "salvo_no_historico_cache" else 500
        return response, status_code

    except Exception as e:
        print("Erro no teste de cache LeaderTrack:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/salvar-registro-semana-leadertrack", methods=["POST", "OPTIONS"])
def salvar_registro_semana_leadertrack():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        empresa = str(dados.get("empresa") or "").strip().lower()
        codrodada = str(dados.get("codrodada") or "").strip().lower()
        email_lider = str(dados.get("emailLider") or "").strip().lower()
        nome_lider = str(dados.get("nomeLider") or "").strip()
        contexto = str(dados.get("contexto") or "").strip()
        equipe_tipo = str(dados.get("equipeTipo") or dados.get("tipoEquipe") or "direta").strip().lower()
        semana = int(dados.get("semana") or 0)
        source_key = str(dados.get("sourceKey") or dados.get("gapId") or dados.get("grupoId") or "").strip()
        feedback = dados.get("feedback") if isinstance(dados.get("feedback"), dict) else {}
        semana_payload = dados.get("semanaPayload") if isinstance(dados.get("semanaPayload"), dict) else {}
        gerado_por = dados.get("geradoPor")
        contexto_ids = {
            "cliente_id": dados.get("cliente_id") or dados.get("clienteId"),
            "holding_id": dados.get("holding_id") or dados.get("holdingId"),
            "empresa_id": dados.get("empresa_id") or dados.get("empresaId"),
            "filial_id": dados.get("filial_id") or dados.get("filialId"),
        }

        if not empresa or not codrodada or not email_lider or not source_key or semana < 1:
            response = jsonify({
                "erro": "Campos obrigatorios ausentes para salvar registro da semana.",
                "campos_necessarios": ["empresa", "codrodada", "emailLider", "sourceKey", "semana"]
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        if not any(str(value or "").strip() for value in feedback.values()):
            response = jsonify({
                "erro": "Registro vazio.",
                "orientacao": "Preencha pelo menos um campo antes de salvar."
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        cache_key = leadertrack_cache_key(
            empresa=empresa,
            codrodada=codrodada,
            email_lider=email_lider,
            equipe_tipo=equipe_tipo,
            gap_id=source_key,
            etapa="registro_semana",
            semana_inicio=semana,
            semana_fim=semana,
        )
        evento = {
            "cliente_id": contexto_ids.get("cliente_id"),
            "holding_id": contexto_ids.get("holding_id"),
            "empresa_id": contexto_ids.get("empresa_id"),
            "filial_id": contexto_ids.get("filial_id"),
            "profissional_email": email_lider,
            "profissional_nome": nome_lider,
            "empresa": empresa,
            "contexto": contexto,
            "origem": "leadertrack_pdi_registro_semanal",
            "tipo_evento": "registro_semana_pdi",
            "descricao_evento": f"Registro do lider salvo para a semana {semana} do PDI LeaderTrack.",
            "dados_antes": {
                "source_key": source_key,
                "semana": semana,
                "semana_payload_resumo": {
                    "foco_da_semana": semana_payload.get("foco_da_semana"),
                    "objetivo": semana_payload.get("objetivo"),
                    "indicador": semana_payload.get("indicador"),
                },
            },
            "dados_depois": {
                "cache_key": cache_key,
                "source_key": source_key,
                "semana": semana,
                "registro_do_lider": feedback,
            },
            "registrado_por": gerado_por,
        }
        saved = supabase_insert("leadertrack_pdi_historico", evento)
        response = jsonify({
            "status": "ok",
            "historico_id": saved.get("id") if isinstance(saved, dict) else None,
            "cache_key": cache_key,
            "mensagem": "Registro da semana salvo no historico LeaderTrack.",
        })
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro ao salvar registro semanal LeaderTrack:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/listar-registros-semana-leadertrack", methods=["POST", "OPTIONS"])
def listar_registros_semana_leadertrack():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        empresa = str(dados.get("empresa") or "").strip().lower()
        codrodada = str(dados.get("codrodada") or "").strip().lower()
        email_lider = str(dados.get("emailLider") or "").strip().lower()

        if not empresa or not codrodada or not email_lider:
            response = jsonify({
                "erro": "Campos obrigatorios ausentes.",
                "campos_necessarios": ["empresa", "codrodada", "emailLider"]
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        url = f"{SUPABASE_REST_URL}/leadertrack_pdi_historico"
        params = {
            "select": "id,dados_depois,data_evento,registrado_por",
            "empresa": f"eq.{empresa}",
            "profissional_email": f"eq.{email_lider}",
            "origem": "eq.leadertrack_pdi_registro_semanal",
            "tipo_evento": "eq.registro_semana_pdi",
            "order": "data_evento.desc",
            "limit": 300,
        }
        response_db = requests.get(
            url,
            headers=supabase_headers(prefer_return=False, use_service_role=True),
            params=params,
            timeout=60,
        )
        if response_db.status_code >= 300:
            raise RuntimeError(f"Erro ao listar registros semanais: HTTP {response_db.status_code} - {response_db.text}")

        registros = {}
        linhas = []
        for row in response_db.json() or []:
            dados_depois = row.get("dados_depois") or {}
            if isinstance(dados_depois, str):
                try:
                    dados_depois = json.loads(dados_depois)
                except Exception:
                    dados_depois = {}
            cache_key = str(dados_depois.get("cache_key") or "")
            if f"|{codrodada}|" not in cache_key:
                continue
            source_key = str(dados_depois.get("source_key") or "").strip()
            semana = dados_depois.get("semana")
            feedback = dados_depois.get("registro_do_lider") if isinstance(dados_depois.get("registro_do_lider"), dict) else {}
            if not source_key or not semana or not feedback:
                continue
            key = f"{source_key}:{semana}"
            if key not in registros:
                registros[key] = feedback
                linhas.append({
                    "key": key,
                    "source_key": source_key,
                    "semana": semana,
                    "feedback": feedback,
                    "historico_id": row.get("id"),
                    "data_evento": row.get("data_evento"),
                    "registrado_por": row.get("registrado_por"),
                })

        response = jsonify({
            "status": "ok",
            "total": len(linhas),
            "registros": registros,
            "linhas": linhas,
        })
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro ao listar registros semanais LeaderTrack:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/gerar-pdi-leadertrack-afirmacao", methods=["POST", "OPTIONS"])
def gerar_pdi_leadertrack_afirmacao():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'CORS preflight OK'})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
        response.headers["Access-Control-Allow-Methods"] = "POST,OPTIONS"
        return response

    try:
        dados = request.get_json() or {}
        empresa = dados.get("empresa", "").lower()
        codrodada = dados.get("codrodada", "").lower()
        email_lider = dados.get("emailLider", "").lower()
        nome_lider = dados.get("nomeLider", "")
        contexto = dados.get("contexto", "")
        equipe_tipo = str(dados.get("equipeTipo") or dados.get("tipoEquipe") or "direta").strip().lower()
        etapa = str(dados.get("etapa") or "diagnostico").strip().lower()
        indicadores_disponiveis = dados.get("indicadoresOperacionaisDisponiveis", [])
        usar_cache = bool_param(dados.get("usarCache"), True)
        gravar_cache = bool_param(dados.get("gravarCache"), True)
        gerado_por = dados.get("geradoPor")

        contexto_ids = {
            "cliente_id": dados.get("cliente_id") or dados.get("clienteId"),
            "holding_id": dados.get("holding_id") or dados.get("holdingId"),
            "empresa_id": dados.get("empresa_id") or dados.get("empresaId"),
            "filial_id": dados.get("filial_id") or dados.get("filialId"),
        }

        if not empresa or not codrodada or not email_lider:
            response = jsonify({
                "erro": "Campos obrigatorios ausentes.",
                "campos_necessarios": ["empresa", "codrodada", "emailLider"]
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 400

        grupo_enviado = dados.get("grupo") if isinstance(dados.get("grupo"), dict) else {}
        grupo_id_previo = str(
            dados.get("grupoId")
            or dados.get("grupo_id")
            or grupo_enviado.get("grupo_id")
            or ""
        ).strip()
        gap_id_previo = str(dados.get("gapId") or dados.get("gap_id") or "").strip()
        cache_id_previo = grupo_id_previo if etapa == "integrado" else gap_id_previo
        if usar_cache and cache_id_previo:
            intervalo_previo = None
            if etapa != "diagnostico":
                intervalo_previo = intervalo_etapa_leadertrack(etapa, dados)
            if etapa == "diagnostico":
                cache_key_previa = leadertrack_cache_key(
                    empresa, codrodada, email_lider, equipe_tipo, cache_id_previo, etapa
                )
                cached_payload = buscar_cache_leadertrack(empresa, email_lider, cache_key_previa)
                if cached_payload:
                    response = jsonify(cached_payload)
                    response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
                    return response, 200
            elif etapa == "integrado":
                inicio_previo, fim_previo = intervalo_previo or (1, 4)
                cache_key_previa = leadertrack_cache_key(
                    empresa, codrodada, email_lider, equipe_tipo, cache_id_previo, etapa, inicio_previo, fim_previo
                )
                cached_payload = buscar_cache_leadertrack(empresa, email_lider, cache_key_previa)
                if cached_payload:
                    response = jsonify(cached_payload)
                    response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
                    return response, 200
            elif intervalo_previo:
                inicio_previo, fim_previo = intervalo_previo
                cache_key_previa = leadertrack_cache_key(
                    empresa, codrodada, email_lider, equipe_tipo, cache_id_previo, etapa, inicio_previo, fim_previo
                )
                cached_payload = buscar_cache_leadertrack(empresa, email_lider, cache_key_previa)
                if cached_payload:
                    response = jsonify(cached_payload)
                    response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
                    return response, 200

        dados_arquetipos_comparativo = buscar_json_supabase(
            "arquetipos_grafico_comparativo", empresa, codrodada, email_lider
        )
        dados_arquetipos_analitico = buscar_json_supabase(
            "arquetipos_analitico", empresa, codrodada, email_lider
        )
        guia_arquetipos = buscar_json_supabase(
            "arquetipos_parecer_ia", empresa, codrodada, email_lider
        )
        dados_microambiente_analitico = buscar_json_microambiente(
            "microambiente_analitico", empresa, codrodada, email_lider
        )
        dados_microambiente_subdimensao = buscar_json_microambiente(
            "microambiente_grafico_mediaequipe_subdimensao", empresa, codrodada, email_lider
        )
        dados_microambiente_termometro_gaps = buscar_json_microambiente(
            "microambiente_termometro_gaps", empresa, codrodada, email_lider
        )
        dados_microambiente_waterfall_gaps = buscar_json_microambiente(
            "microambiente_waterfall_gaps", empresa, codrodada, email_lider
        )
        guia_microambiente = buscar_json_microambiente(
            "microambiente_parecer_ia", empresa, codrodada, email_lider
        )

        if not dados_arquetipos_comparativo or not dados_microambiente_analitico:
            response = jsonify({
                "erro": "Dados LeaderTrack insuficientes para gerar PDI da afirmacao.",
                "dados_encontrados": {
                    "arquetipos_grafico_comparativo": bool(dados_arquetipos_comparativo),
                    "microambiente_analitico": bool(dados_microambiente_analitico),
                }
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 404

        arquetipos = archetype_summary(dados_arquetipos_comparativo)
        todas_afirmacoes = microenvironment_affirmations(dados_microambiente_analitico)
        gap = None if etapa == "integrado" else selecionar_gap_leadertrack(todas_afirmacoes, dados)
        if etapa != "integrado" and not gap:
            response = jsonify({
                "erro": "Afirmacao/gap nao encontrado.",
                "orientacao": "Informe gapId, questao ou envie o objeto gap retornado pela devolutiva estruturada."
            })
            response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
            return response, 404

        leader = {
            "nome": nome_lider,
            "email": email_lider,
            "rodada": codrodada,
            "contexto": contexto,
            "equipe_tipo": equipe_tipo,
            "contexto_ids": contexto_ids,
        }
        prompt_base = carregar_prompt_leadertrack()
        gap_id = leadertrack_gap_id(gap) if gap else grupo_id_previo

        if etapa == "diagnostico":
            cache_key = leadertrack_cache_key(empresa, codrodada, email_lider, equipe_tipo, gap_id, etapa)
            if usar_cache:
                cached_payload = buscar_cache_leadertrack(empresa, email_lider, cache_key)
                if cached_payload:
                    response = jsonify(cached_payload)
                    response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
                    return response, 200

            prompt = build_diagnostic_prompt(
                leader=leader,
                arquetipos=arquetipos,
                gap=gap,
                indicadores_disponiveis=indicadores_disponiveis,
            )
            prompt_base_semanal = (
                "Voce e o LeaderTrackbot. Use exclusivamente o CONTEXTO_JSON enviado pelo usuario. "
                "Nao invente dados, nomes, indicadores numericos ou teorias externas. "
                "Gere um PDI consultivo, pratico e fiel ao gap, aos arquetipos e ao microambiente informados. "
                "Responda somente JSON valido no formato solicitado."
            )
            resposta_ia = gerar_resposta_ia_leadertrack_enxuta(
                pergunta=prompt,
                prompt_base=prompt_base_semanal,
            )
            resultado = parse_json_response(resposta_ia)
            payload = {
                "status": "ok",
                "etapa": etapa,
                "gap_id": gap_id,
                "gap": gap,
                "diagnostico": resultado,
                "persistencia": "nao_salvo",
                "fonte": "ia",
                "cache": {
                    "status": "miss",
                    "cache_key": cache_key,
                },
                "proxima_etapa_sugerida": "semanas_1_4",
            }
            if gravar_cache:
                persistencia_cache = salvar_cache_leadertrack(
                    empresa=empresa,
                    contexto=contexto,
                    contexto_ids=contexto_ids,
                    email_lider=email_lider,
                    nome_lider=nome_lider,
                    cache_key=cache_key,
                    payload=payload,
                    gerado_por=gerado_por,
                )
                payload["persistencia"] = persistencia_cache
                if persistencia_cache.get("status") == "salvo_no_historico_cache":
                    payload["cache"]["status"] = "saved"
                    payload["cache"]["historico_id"] = persistencia_cache.get("historico_id")
                else:
                    payload["cache"]["status"] = "save_failed"
                    payload["cache"]["erro"] = persistencia_cache.get("erro")
        elif etapa == "integrado":
            grupo = grupo_enviado or {}
            grupo_id = str(grupo.get("grupo_id") or grupo_id_previo or "").strip()
            if not grupo_id or not isinstance(grupo.get("afirmacoes"), list) or not grupo.get("afirmacoes"):
                response = jsonify({
                    "erro": "Grupo tematico ausente ou incompleto.",
                    "orientacao": "Envie grupoId e o objeto grupo retornado em agrupamentos_tematicos."
                })
                response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
                return response, 400

            inicio, fim = intervalo_etapa_leadertrack(etapa, dados) or (1, 4)
            cache_key = leadertrack_cache_key(empresa, codrodada, email_lider, equipe_tipo, grupo_id, etapa, inicio, fim)
            if usar_cache:
                cached_payload = buscar_cache_leadertrack(empresa, email_lider, cache_key)
                if cached_payload:
                    response = jsonify(cached_payload)
                    response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
                    return response, 200

            if inicio == 1 and fim == 1:
                prompt = build_integrated_week1_prompt(
                    leader=leader,
                    arquetipos=arquetipos,
                    group=grupo,
                    indicadores_disponiveis=indicadores_disponiveis,
                )
                max_tokens = 2800
                timeout = 25
            else:
                prompt = build_integrated_plan_prompt(
                    leader=leader,
                    arquetipos=arquetipos,
                    group=grupo,
                    indicadores_disponiveis=indicadores_disponiveis,
                    start_week=inicio,
                    end_week=fim,
                    feedback_semanas=dados.get("feedbackSemanas") or dados.get("feedback_semanas") or {},
                )
                max_tokens = 2400 if inicio == fim else 4200
                timeout = 25 if inicio == fim else 45

            resposta_ia = gerar_resposta_ia_leadertrack_enxuta(
                pergunta=prompt,
                prompt_base=prompt_base,
                model="gpt-4.1-mini",
                max_tokens=max_tokens,
                timeout=timeout,
                temperature=0.35,
            )
            resultado = parse_json_response(resposta_ia)
            resultado = normalizar_plano_semanal_leadertrack(resultado)
            resultado = revisar_plano_leadertrack_se_incompleto(resultado, prompt_base)
            payload = {
                "status": "ok",
                "etapa": etapa,
                "grupo_id": grupo_id,
                "grupo": grupo,
                "semana_inicio": inicio,
                "semana_fim": fim,
                "plano_integrado": resultado,
                "persistencia": "nao_salvo",
                "fonte": "ia",
                "cache": {
                    "status": "miss",
                    "cache_key": cache_key,
                },
                "proxima_etapa_sugerida": None,
                "observacao": "Plano integrado gerado para uso consultivo. Ainda nao enviado ao PDI oficial.",
            }
            if gravar_cache:
                persistencia_cache = salvar_cache_leadertrack(
                    empresa=empresa,
                    contexto=contexto,
                    contexto_ids=contexto_ids,
                    email_lider=email_lider,
                    nome_lider=nome_lider,
                    cache_key=cache_key,
                    payload=payload,
                    gerado_por=gerado_por,
                )
                payload["persistencia"] = persistencia_cache
                if persistencia_cache.get("status") == "salvo_no_historico_cache":
                    payload["cache"]["status"] = "saved"
                    payload["cache"]["historico_id"] = persistencia_cache.get("historico_id")
                else:
                    payload["cache"]["status"] = "save_failed"
                    payload["cache"]["erro"] = persistencia_cache.get("erro")
        else:
            intervalo = intervalo_etapa_leadertrack(etapa, dados)
            if not intervalo:
                response = jsonify({
                    "erro": "Etapa invalida.",
                    "etapas_validas": ["diagnostico", "integrado", "semanas_1_4", "semanas_5_8", "semanas_9_12"]
                })
                response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
                return response, 400

            diagnostico = dados.get("diagnostico") or {
                "status": "diagnostico_nao_enviado",
                "orientacao": "Use preferencialmente o diagnostico gerado na etapa diagnostico como entrada desta chamada.",
            }
            inicio, fim = intervalo
            cache_key = leadertrack_cache_key(empresa, codrodada, email_lider, equipe_tipo, gap_id, etapa, inicio, fim)
            if usar_cache:
                cached_payload = buscar_cache_leadertrack(empresa, email_lider, cache_key)
                if cached_payload:
                    response = jsonify(cached_payload)
                    response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
                    return response, 200

            prompt = build_weekly_prompt(
                leader=leader,
                arquetipos=arquetipos,
                gap=gap,
                diagnostic=diagnostico,
                start_week=inicio,
                end_week=fim,
                indicadores_disponiveis=indicadores_disponiveis,
            )
            prompt_base_semanal = (
                "Voce e o LeaderTrackbot. Use exclusivamente o CONTEXTO_JSON enviado pelo usuario. "
                "Nao invente dados, nomes, indicadores numericos ou teorias externas. "
                "Gere um PDI consultivo, pratico e fiel ao gap, aos arquetipos e ao microambiente informados. "
                "Responda somente JSON valido no formato solicitado."
            )
            resposta_ia = gerar_resposta_ia_leadertrack_enxuta(
                pergunta=prompt,
                prompt_base=prompt_base_semanal,
                model="gpt-4.1-mini",
                max_tokens=3200,
                timeout=40,
                temperature=0.35,
            )
            resultado = parse_json_response(resposta_ia)
            resultado = normalizar_plano_semanal_leadertrack(resultado)
            resultado = revisar_plano_leadertrack_se_incompleto(resultado, prompt_base)
            payload = {
                "status": "ok",
                "etapa": etapa,
                "gap_id": gap_id,
                "gap": gap,
                "diagnostico_usado": diagnostico,
                "plano_parcial": resultado,
                "persistencia": "nao_salvo",
                "fonte": "ia",
                "cache": {
                    "status": "miss",
                    "cache_key": cache_key,
                },
                "proxima_etapa_sugerida": None,
            }
            if gravar_cache:
                persistencia_cache = salvar_cache_leadertrack(
                    empresa=empresa,
                    contexto=contexto,
                    contexto_ids=contexto_ids,
                    email_lider=email_lider,
                    nome_lider=nome_lider,
                    cache_key=cache_key,
                    payload=payload,
                    gerado_por=gerado_por,
                )
                payload["persistencia"] = persistencia_cache
                if persistencia_cache.get("status") == "salvo_no_historico_cache":
                    payload["cache"]["status"] = "saved"
                    payload["cache"]["historico_id"] = persistencia_cache.get("historico_id")
                else:
                    payload["cache"]["status"] = "save_failed"
                    payload["cache"]["erro"] = persistencia_cache.get("erro")

        response = jsonify(payload)
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 200

    except Exception as e:
        print("Erro ao gerar PDI LeaderTrack por afirmacao:", e)
        response = jsonify({"erro": str(e)})
        response.headers["Access-Control-Allow-Origin"] = "https://gestor.thehrkey.tech"
        return response, 500


@app.route("/teste-chat-leadertrack", methods=["GET"])
def teste_chat_leadertrack_get():
    try:
        empresa = request.args.get("empresa", "").lower()
        codrodada = request.args.get("codrodada", "").lower()
        email_lider = request.args.get("emailLider", "").lower()
        pergunta = request.args.get(
            "pergunta",
            "Como posso melhorar o microambiente da minha equipe com base nos meus arquétipos?"
        )

        if not empresa or not codrodada or not email_lider:
            return jsonify({
                "status": "erro",
                "mensagem": "Informe empresa, codrodada e emailLider na URL.",
                "exemplo": "/teste-chat-leadertrack?empresa=empresa_teste&codrodada=rodada1&emailLider=lider@teste.com"
            }), 400

        prompt_base = carregar_prompt_leadertrack()

        dados_arquetipos_comparativo = buscar_json_supabase(
            "arquetipos_grafico_comparativo",
            empresa,
            codrodada,
            email_lider
        )

        dados_arquetipos_analitico = buscar_json_supabase(
            "arquetipos_analitico",
            empresa,
            codrodada,
            email_lider
        )

        guia_arquetipos = buscar_json_supabase(
            "arquetipos_parecer_ia",
            empresa,
            codrodada,
            email_lider
        )

        dados_microambiente_analitico = buscar_json_microambiente(
            "microambiente_analitico",
            empresa,
            codrodada,
            email_lider
        )

        dados_microambiente_subdimensao = buscar_json_microambiente(
            "microambiente_grafico_mediaequipe_subdimensao",
            empresa,
            codrodada,
            email_lider
        )

        dados_microambiente_termometro_gaps = buscar_json_microambiente(
            "microambiente_termometro_gaps",
            empresa,
            codrodada,
            email_lider
        )

        dados_microambiente_waterfall_gaps = buscar_json_microambiente(
            "microambiente_waterfall_gaps",
            empresa,
            codrodada,
            email_lider
        )

        guia_microambiente = buscar_json_microambiente(
            "microambiente_parecer_ia",
            empresa,
            codrodada,
            email_lider
        )

        

        return jsonify({
            "status": "ok",
            "pergunta": pergunta,
            "prompt_carregado": True,
            "tamanho_prompt": len(prompt_base),
            "dados_encontrados": {
                "arquetipos_grafico_comparativo": "ENCONTRADO" if dados_arquetipos_comparativo else "NÃO ENCONTRADO",
                "arquetipos_analitico": "ENCONTRADO" if dados_arquetipos_analitico else "NÃO ENCONTRADO",
                "arquetipos_parecer_ia_guia": "ENCONTRADO" if guia_arquetipos else "NÃO ENCONTRADO",

                "microambiente_analitico": "ENCONTRADO" if dados_microambiente_analitico else "NÃO ENCONTRADO",
                "microambiente_grafico_mediaequipe_subdimensao": "ENCONTRADO" if dados_microambiente_subdimensao else "NÃO ENCONTRADO",
                "microambiente_termometro_gaps": "ENCONTRADO" if dados_microambiente_termometro_gaps else "NÃO ENCONTRADO",
                "microambiente_waterfall_gaps": "ENCONTRADO" if dados_microambiente_waterfall_gaps else "NÃO ENCONTRADO",
                "microambiente_parecer_ia_guia": "ENCONTRADO" if guia_microambiente else "NÃO ENCONTRADO",

                
            }
        }), 200

    except Exception as e:
        print("Erro no teste chat Leadertrack GET:", e)
        return jsonify({
            "status": "erro",
            "mensagem": str(e)
        }), 500



@app.route("/teste-ia-leadertrack", methods=["GET"])
def teste_ia_leadertrack_get():
    try:
        empresa = request.args.get("empresa", "").lower()
        codrodada = request.args.get("codrodada", "").lower()
        email_lider = request.args.get("emailLider", "").lower()
        pergunta = request.args.get(
            "pergunta",
            "Como posso melhorar o microambiente da minha equipe com base nos meus arquétipos?"
        )

        if not empresa or not codrodada or not email_lider:
            return jsonify({
                "status": "erro",
                "mensagem": "Informe empresa, codrodada e emailLider na URL.",
                "exemplo": "/teste-ia-leadertrack?empresa=fastco&codrodada=av1225&emailLider=felipe@astro34.com.br"
            }), 400

        prompt_base = carregar_prompt_leadertrack()

        dados_arquetipos_comparativo = buscar_json_supabase(
            "arquetipos_grafico_comparativo",
            empresa,
            codrodada,
            email_lider
        )

        dados_arquetipos_analitico = buscar_json_supabase(
            "arquetipos_analitico",
            empresa,
            codrodada,
            email_lider
        )

        guia_arquetipos = buscar_json_supabase(
            "arquetipos_parecer_ia",
            empresa,
            codrodada,
            email_lider
        )

        dados_microambiente_analitico = buscar_json_microambiente(
            "microambiente_analitico",
            empresa,
            codrodada,
            email_lider
        )

        dados_microambiente_subdimensao = buscar_json_microambiente(
            "microambiente_grafico_mediaequipe_subdimensao",
            empresa,
            codrodada,
            email_lider
        )

        dados_microambiente_termometro_gaps = buscar_json_microambiente(
            "microambiente_termometro_gaps",
            empresa,
            codrodada,
            email_lider
        )

        dados_microambiente_waterfall_gaps = buscar_json_microambiente(
            "microambiente_waterfall_gaps",
            empresa,
            codrodada,
            email_lider
        )

        guia_microambiente = buscar_json_microambiente(
            "microambiente_parecer_ia",
            empresa,
            codrodada,
            email_lider
        )

        resposta_ia = gerar_resposta_ia_leadertrack(
            pergunta=pergunta,
            prompt_base=prompt_base,
            empresa=empresa,
            codrodada=codrodada,
            email_lider=email_lider,
            pagina_atual="/teste-ia-leadertrack",
            url_atual=request.url, 
            dados_arquetipos_comparativo=dados_arquetipos_comparativo,
            dados_arquetipos_analitico=dados_arquetipos_analitico,
            guia_arquetipos=guia_arquetipos,
            dados_microambiente_analitico=dados_microambiente_analitico,
            dados_microambiente_subdimensao=dados_microambiente_subdimensao,
            dados_microambiente_termometro_gaps=dados_microambiente_termometro_gaps,
            dados_microambiente_waterfall_gaps=dados_microambiente_waterfall_gaps,
            guia_microambiente=guia_microambiente
        )

        return jsonify({
            "status": "ok",
            "pergunta": pergunta,
            "resposta": resposta_ia
        }), 200

    except Exception as e:
        print("Erro no teste IA Leadertrack GET:", e)
        return jsonify({
            "status": "erro",
            "mensagem": str(e)
        }), 500




if __name__ == "__main__":
    app.run(debug=True)
